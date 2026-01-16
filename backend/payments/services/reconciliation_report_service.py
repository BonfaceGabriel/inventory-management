"""
Enhanced Reconciliation Report Service for XLSX Generation

Generates comprehensive XLSX reports with:
- Separate sheets per gateway
- Status filtering per sheet
- Reduced field set (only required columns)
- Gateway name resolution for manual transactions
- Date in filename
"""

from decimal import Decimal
from datetime import date, datetime
from io import BytesIO
from typing import Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from django.utils import timezone
from django.db.models import QuerySet, Q
import logging

from payments.models import Transaction, PaymentGateway, ManualPayment

logger = logging.getLogger(__name__)


class ReconciliationReportService:
    """
    Enhanced service for generating reconciliation reports in XLSX format.

    Key features:
    - One sheet per gateway
    - Status filters
    - Minimal field set
    - Manual payment gateway resolution
    - Date-stamped filenames
    """

    # Required columns only
    HEADERS = [
        'Transaction ID',
        'Timestamp',
        'Amount (KES)',
        'Amount Fulfilled (KES)',
        'Amount Remaining (KES)',
        'Sender Phone',
        'Gateway Name',
        'Status',
    ]

    @staticmethod
    def generate_daily_report_xlsx(report_date: date = None) -> tuple[BytesIO, str]:
        """
        Generate XLSX report for a specific date with gateway sheets.

        Args:
            report_date: Date to generate report for (defaults to today)

        Returns:
            Tuple of (BytesIO object, filename)
        """
        if report_date is None:
            report_date = timezone.now().date()

        logger.info(f"Generating reconciliation XLSX report for {report_date}")

        # Get all transactions for the date
        start_datetime = timezone.make_aware(datetime.combine(report_date, datetime.min.time()))
        end_datetime = timezone.make_aware(datetime.combine(report_date, datetime.max.time()))

        # IMPORTANT: Exclude COMBINED_FULFILLED transactions to avoid double-counting revenue
        # When transactions are combined, children are marked COMBINED_FULFILLED and the parent
        # CombinedOrder tracks the total. Including both would duplicate revenue figures.
        transactions = Transaction.objects.filter(
            timestamp__gte=start_datetime,
            timestamp__lte=end_datetime
        ).exclude(
            status=Transaction.OrderStatus.COMBINED_FULFILLED
        ).select_related('gateway').prefetch_related('manual_payments')

        # Generate filename with date
        filename = f"Reconciliation_Report_{report_date.strftime('%Y%m%d')}_Generated_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        # Create workbook
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # Get all gateways (including manual payments)
        gateway_ids = set()
        for txn in transactions:
            if txn.gateway_id:
                gateway_ids.add(txn.gateway_id)
            elif txn.manual_payments.exists():
                # For manual payments, try to infer gateway from payment method
                gateway_ids.add(None)  # None represents manual payments

        gateways = PaymentGateway.objects.filter(id__in=[gid for gid in gateway_ids if gid is not None])

        # Create a sheet for each gateway
        for gateway in gateways:
            ReconciliationReportService._create_gateway_sheet(
                wb,
                gateway,
                transactions.filter(gateway=gateway),
                report_date
            )

        # Create sheet for manual payments (no gateway)
        manual_txns = transactions.filter(gateway__isnull=True)
        if manual_txns.exists():
            ReconciliationReportService._create_manual_payments_sheet(
                wb,
                manual_txns,
                report_date
            )

        # Save to buffer
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        logger.info(f"Successfully generated reconciliation XLSX for {report_date}")
        return output, filename

    @staticmethod
    def generate_date_range_report_xlsx(start_date: date, end_date: date) -> tuple[BytesIO, str]:
        """
        Generate XLSX report for a date range with gateway sheets.

        Args:
            start_date: Start date
            end_date: End date

        Returns:
            Tuple of (BytesIO object, filename)
        """
        logger.info(f"Generating reconciliation XLSX report from {start_date} to {end_date}")

        # Get all transactions for the date range
        start_datetime = timezone.make_aware(datetime.combine(start_date, datetime.min.time()))
        end_datetime = timezone.make_aware(datetime.combine(end_date, datetime.max.time()))

        # IMPORTANT: Exclude COMBINED_FULFILLED transactions to avoid double-counting revenue
        transactions = Transaction.objects.filter(
            timestamp__gte=start_datetime,
            timestamp__lte=end_datetime
        ).exclude(
            status=Transaction.OrderStatus.COMBINED_FULFILLED
        ).select_related('gateway').prefetch_related('manual_payments')

        # Generate filename with date range
        filename = f"Reconciliation_Report_{start_date.strftime('%Y%m%d')}_to_{end_date.strftime('%Y%m%d')}_Generated_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        # Create workbook
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # Get all gateways
        gateway_ids = set()
        for txn in transactions:
            if txn.gateway_id:
                gateway_ids.add(txn.gateway_id)
            elif txn.manual_payments.exists():
                gateway_ids.add(None)

        gateways = PaymentGateway.objects.filter(id__in=[gid for gid in gateway_ids if gid is not None])

        # Create a sheet for each gateway
        for gateway in gateways:
            ReconciliationReportService._create_gateway_sheet(
                wb,
                gateway,
                transactions.filter(gateway=gateway),
                None  # No specific date for range reports
            )

        # Create sheet for manual payments
        manual_txns = transactions.filter(gateway__isnull=True)
        if manual_txns.exists():
            ReconciliationReportService._create_manual_payments_sheet(
                wb,
                manual_txns,
                None
            )

        # Save to buffer
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        logger.info(f"Successfully generated reconciliation XLSX from {start_date} to {end_date}")
        return output, filename

    @staticmethod
    def _create_gateway_sheet(wb: Workbook, gateway: PaymentGateway, transactions: QuerySet, report_date: Optional[date]):
        """
        Create a sheet for a specific gateway with status sections.

        Args:
            wb: Workbook object
            gateway: Gateway object
            transactions: QuerySet of transactions for this gateway
            report_date: Optional report date (for title)
        """
        # Sanitize sheet name (Excel has character limits and restrictions)
        sheet_name = gateway.name[:30]  # Excel max 31 chars, leave room for safety
        ws = wb.create_sheet(title=sheet_name)

        # Define styles
        header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4A5568', end_color='4A5568', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        section_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
        section_fill = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')

        cell_alignment = Alignment(horizontal='left', vertical='center')
        number_alignment = Alignment(horizontal='right', vertical='center')
        center_alignment = Alignment(horizontal='center', vertical='center')

        border = Border(
            left=Side(style='thin', color='CBD5E0'),
            right=Side(style='thin', color='CBD5E0'),
            top=Side(style='thin', color='CBD5E0'),
            bottom=Side(style='thin', color='CBD5E0')
        )

        # Set column widths
        ws.column_dimensions['A'].width = 15  # Transaction ID
        ws.column_dimensions['B'].width = 20  # Timestamp
        ws.column_dimensions['C'].width = 15  # Amount
        ws.column_dimensions['D'].width = 18  # Amount Fulfilled
        ws.column_dimensions['E'].width = 18  # Amount Remaining
        ws.column_dimensions['F'].width = 15  # Sender Phone
        ws.column_dimensions['G'].width = 20  # Gateway Name
        ws.column_dimensions['H'].width = 18  # Status

        row_num = 1

        # Group transactions by status
        statuses = [
            ('NOT_PROCESSED', 'Not Processed'),
            ('PROCESSING', 'Processing'),
            ('PARTIALLY_FULFILLED', 'Partially Fulfilled'),
            ('FULFILLED', 'Fulfilled'),
            ('CANCELLED', 'Cancelled'),
        ]

        for status_code, status_label in statuses:
            status_txns = transactions.filter(status=status_code)

            if not status_txns.exists():
                continue

            # Section header
            section_cell = ws.cell(row=row_num, column=1, value=f"{status_label} ({status_txns.count()})")
            section_cell.font = section_font
            section_cell.fill = section_fill
            section_cell.alignment = center_alignment

            # Merge cells for section header
            ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=len(ReconciliationReportService.HEADERS))
            for col in range(1, len(ReconciliationReportService.HEADERS) + 1):
                ws.cell(row=row_num, column=col).border = border

            row_num += 1

            # Column headers
            for col_num, header in enumerate(ReconciliationReportService.HEADERS, 1):
                cell = ws.cell(row=row_num, column=col_num, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border

            row_num += 1

            # Data rows
            section_total_amount = Decimal('0.00')
            section_total_fulfilled = Decimal('0.00')
            section_total_remaining = Decimal('0.00')

            for txn in status_txns:
                # Transaction ID
                cell = ws.cell(row=row_num, column=1, value=txn.tx_id or '')
                cell.alignment = cell_alignment
                cell.border = border

                # Timestamp
                cell = ws.cell(row=row_num, column=2, value=txn.timestamp.strftime('%Y-%m-%d %H:%M:%S') if txn.timestamp else '')
                cell.alignment = center_alignment
                cell.border = border

                # Amount
                cell = ws.cell(row=row_num, column=3, value=float(txn.amount))
                cell.alignment = number_alignment
                cell.number_format = '#,##0.00'
                cell.border = border
                section_total_amount += txn.amount

                # Amount Fulfilled
                cell = ws.cell(row=row_num, column=4, value=float(txn.amount_paid))
                cell.alignment = number_alignment
                cell.number_format = '#,##0.00'
                cell.border = border
                section_total_fulfilled += txn.amount_paid

                # Amount Remaining
                cell = ws.cell(row=row_num, column=5, value=float(txn.remaining_amount))
                cell.alignment = number_alignment
                cell.number_format = '#,##0.00'
                cell.border = border
                section_total_remaining += txn.remaining_amount

                # Sender Phone
                cell = ws.cell(row=row_num, column=6, value=txn.sender_phone or '')
                cell.alignment = cell_alignment
                cell.border = border

                # Gateway Name
                cell = ws.cell(row=row_num, column=7, value=gateway.name)
                cell.alignment = cell_alignment
                cell.border = border

                # Status
                cell = ws.cell(row=row_num, column=8, value=txn.get_status_display())
                cell.alignment = center_alignment
                cell.border = border

                row_num += 1

            # Section subtotal
            subtotal_font = Font(name='Calibri', size=10, bold=True)
            subtotal_fill = PatternFill(start_color='E2E8F0', end_color='E2E8F0', fill_type='solid')

            cell = ws.cell(row=row_num, column=2, value=f'{status_label} Subtotal')
            cell.font = subtotal_font
            cell.fill = subtotal_fill
            cell.alignment = center_alignment
            cell.border = border

            cell = ws.cell(row=row_num, column=3, value=float(section_total_amount))
            cell.font = subtotal_font
            cell.fill = subtotal_fill
            cell.alignment = number_alignment
            cell.number_format = '#,##0.00'
            cell.border = border

            cell = ws.cell(row=row_num, column=4, value=float(section_total_fulfilled))
            cell.font = subtotal_font
            cell.fill = subtotal_fill
            cell.alignment = number_alignment
            cell.number_format = '#,##0.00'
            cell.border = border

            cell = ws.cell(row=row_num, column=5, value=float(section_total_remaining))
            cell.font = subtotal_font
            cell.fill = subtotal_fill
            cell.alignment = number_alignment
            cell.number_format = '#,##0.00'
            cell.border = border

            row_num += 2  # Blank row between sections

        # Gateway total
        total_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
        total_fill = PatternFill(start_color='1F2937', end_color='1F2937', fill_type='solid')

        gateway_total_amount = sum(float(txn.amount) for txn in transactions)
        gateway_total_fulfilled = sum(float(txn.amount_paid) for txn in transactions)
        gateway_total_remaining = sum(float(txn.remaining_amount) for txn in transactions)

        cell = ws.cell(row=row_num, column=2, value='GATEWAY TOTAL')
        cell.font = total_font
        cell.fill = total_fill
        cell.alignment = center_alignment
        cell.border = border

        cell = ws.cell(row=row_num, column=3, value=gateway_total_amount)
        cell.font = total_font
        cell.fill = total_fill
        cell.alignment = number_alignment
        cell.number_format = '#,##0.00'
        cell.border = border

        cell = ws.cell(row=row_num, column=4, value=gateway_total_fulfilled)
        cell.font = total_font
        cell.fill = total_fill
        cell.alignment = number_alignment
        cell.number_format = '#,##0.00'
        cell.border = border

        cell = ws.cell(row=row_num, column=5, value=gateway_total_remaining)
        cell.font = total_font
        cell.fill = total_fill
        cell.alignment = number_alignment
        cell.number_format = '#,##0.00'
        cell.border = border

        # Freeze top row
        ws.freeze_panes = 'A3'

    @staticmethod
    def _create_manual_payments_sheet(wb: Workbook, transactions: QuerySet, report_date: Optional[date]):
        """
        Create a sheet for manual payments (no M-Pesa gateway).

        Args:
            wb: Workbook object
            transactions: QuerySet of manual payment transactions
            report_date: Optional report date
        """
        ws = wb.create_sheet(title="Manual Payments")

        # Same styling as gateway sheets
        header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4A5568', end_color='4A5568', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        section_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
        section_fill = PatternFill(start_color='10B981', end_color='10B981', fill_type='solid')

        cell_alignment = Alignment(horizontal='left', vertical='center')
        number_alignment = Alignment(horizontal='right', vertical='center')
        center_alignment = Alignment(horizontal='center', vertical='center')

        border = Border(
            left=Side(style='thin', color='CBD5E0'),
            right=Side(style='thin', color='CBD5E0'),
            top=Side(style='thin', color='CBD5E0'),
            bottom=Side(style='thin', color='CBD5E0')
        )

        # Set column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 18
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 25  # Wider for payment method
        ws.column_dimensions['H'].width = 18

        row_num = 1

        # Section header
        section_cell = ws.cell(row=row_num, column=1, value=f"Manual Payments ({transactions.count()})")
        section_cell.font = section_font
        section_cell.fill = section_fill
        section_cell.alignment = center_alignment
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=len(ReconciliationReportService.HEADERS))
        for col in range(1, len(ReconciliationReportService.HEADERS) + 1):
            ws.cell(row=row_num, column=col).border = border

        row_num += 1

        # Column headers
        for col_num, header in enumerate(ReconciliationReportService.HEADERS, 1):
            cell = ws.cell(row=row_num, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

        row_num += 1

        # Data rows
        total_amount = Decimal('0.00')
        total_fulfilled = Decimal('0.00')
        total_remaining = Decimal('0.00')

        for txn in transactions:
            # Get manual payment info
            manual_payment = txn.manual_payments.first() if txn.manual_payments.exists() else None
            payment_method = manual_payment.get_payment_method_display() if manual_payment else 'Unknown'

            cell = ws.cell(row=row_num, column=1, value=txn.tx_id or '')
            cell.alignment = cell_alignment
            cell.border = border

            cell = ws.cell(row=row_num, column=2, value=txn.timestamp.strftime('%Y-%m-%d %H:%M:%S') if txn.timestamp else '')
            cell.alignment = center_alignment
            cell.border = border

            cell = ws.cell(row=row_num, column=3, value=float(txn.amount))
            cell.alignment = number_alignment
            cell.number_format = '#,##0.00'
            cell.border = border
            total_amount += txn.amount

            cell = ws.cell(row=row_num, column=4, value=float(txn.amount_paid))
            cell.alignment = number_alignment
            cell.number_format = '#,##0.00'
            cell.border = border
            total_fulfilled += txn.amount_paid

            cell = ws.cell(row=row_num, column=5, value=float(txn.remaining_amount))
            cell.alignment = number_alignment
            cell.number_format = '#,##0.00'
            cell.border = border
            total_remaining += txn.remaining_amount

            cell = ws.cell(row=row_num, column=6, value=txn.sender_phone or '')
            cell.alignment = cell_alignment
            cell.border = border

            # Gateway Name shows payment method for manual payments
            cell = ws.cell(row=row_num, column=7, value=payment_method)
            cell.alignment = cell_alignment
            cell.border = border

            cell = ws.cell(row=row_num, column=8, value=txn.get_status_display())
            cell.alignment = center_alignment
            cell.border = border

            row_num += 1

        # Total
        total_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
        total_fill = PatternFill(start_color='1F2937', end_color='1F2937', fill_type='solid')

        cell = ws.cell(row=row_num, column=2, value='TOTAL')
        cell.font = total_font
        cell.fill = total_fill
        cell.alignment = center_alignment
        cell.border = border

        cell = ws.cell(row=row_num, column=3, value=float(total_amount))
        cell.font = total_font
        cell.fill = total_fill
        cell.alignment = number_alignment
        cell.number_format = '#,##0.00'
        cell.border = border

        cell = ws.cell(row=row_num, column=4, value=float(total_fulfilled))
        cell.font = total_font
        cell.fill = total_fill
        cell.alignment = number_alignment
        cell.number_format = '#,##0.00'
        cell.border = border

        cell = ws.cell(row=row_num, column=5, value=float(total_remaining))
        cell.font = total_font
        cell.fill = total_fill
        cell.alignment = number_alignment
        cell.number_format = '#,##0.00'
        cell.border = border

        ws.freeze_panes = 'A3'

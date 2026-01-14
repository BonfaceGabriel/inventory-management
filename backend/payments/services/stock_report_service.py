"""
Stock Report Service

Generates inventory stock reports with current levels, valuations, and alerts.
Supports historical reporting by reconstructing stock levels from inventory movements.
"""

from decimal import Decimal
from datetime import datetime, date, time
from io import BytesIO
from typing import Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from django.db.models import Sum, F, Q, Value, DecimalField
from django.db.models.functions import Coalesce
from django.utils import timezone
import logging

from payments.models import Product, ProductLine, InventoryMovement, DailyStockReconciliation

logger = logging.getLogger(__name__)


class StockReportService:
    """
    Service for generating inventory stock reports.

    Provides methods to:
    - Generate current stock report (JSON)
    - Generate stock report XLSX
    - Calculate inventory valuations
    - Identify low stock / out of stock items
    """

    @staticmethod
    def generate_stock_report() -> Dict:
        """
        Generate comprehensive stock report with current inventory levels.

        Returns:
            Dictionary with stock summary and product details
        """
        logger.info("Generating stock report")

        # Get all active products with stock info
        products = Product.objects.filter(is_active=True).select_related('product_line').order_by('product_line__name', 'prod_name')

        # Calculate totals using cost_price (buying price)
        total_products = products.count()
        total_stock_value = sum(
            (p.quantity * p.cost_price) for p in products
        )
        total_cost_value = sum(
            (p.quantity * p.cost_price) for p in products
        )

        # Stock status counts
        out_of_stock = products.filter(quantity__lte=0).count()
        low_stock = products.filter(quantity__gt=0, quantity__lte=F('reorder_level')).count()
        in_stock = products.filter(quantity__gt=F('reorder_level')).count()

        # Group by product line
        product_lines_data = []
        product_lines = ProductLine.objects.all().order_by('name')

        for product_line in product_lines:
            line_products = products.filter(product_line=product_line)
            if line_products.exists():
                line_stock_value = sum(
                    (p.quantity * p.cost_price) for p in line_products
                )
                product_lines_data.append({
                    'product_line_name': product_line.name,
                    'product_count': line_products.count(),
                    'total_quantity': sum(p.quantity for p in line_products),
                    'stock_value': float(line_stock_value),
                    'products': [
                        {
                            'id': p.id,
                            'prod_code': p.prod_code,
                            'prod_name': p.prod_name,
                            'sku': p.sku,
                            'barcode': p.barcode,
                            'quantity': p.quantity,
                            'reorder_level': p.reorder_level,
                            'current_price': float(p.current_price),
                            'cost_price': float(p.cost_price),
                            'stock_value': float(p.quantity * p.cost_price),  # Use cost_price
                            'stock_status': StockReportService._get_stock_status(p),
                        }
                        for p in line_products
                    ]
                })

        # Products without product line
        unassigned = products.filter(product_line__isnull=True)
        if unassigned.exists():
            unassigned_stock_value = sum(
                (p.quantity * p.cost_price) for p in unassigned
            )
            product_lines_data.append({
                'product_line_name': 'Unassigned',
                'product_count': unassigned.count(),
                'total_quantity': sum(p.quantity for p in unassigned),
                'stock_value': float(unassigned_stock_value),
                'products': [
                    {
                        'id': p.id,
                        'prod_code': p.prod_code,
                        'prod_name': p.prod_name,
                        'sku': p.sku,
                        'barcode': p.barcode,
                        'quantity': p.quantity,
                        'reorder_level': p.reorder_level,
                        'current_price': float(p.current_price),
                        'cost_price': float(p.cost_price),
                        'stock_value': float(p.quantity * p.cost_price),
                        'stock_status': StockReportService._get_stock_status(p),
                    }
                    for p in unassigned
                ]
            })

        report = {
            'generated_at': timezone.now().isoformat(),
            'summary': {
                'total_products': total_products,
                'total_stock_value': float(total_stock_value),
                'total_cost_value': float(total_cost_value),
                'out_of_stock_count': out_of_stock,
                'low_stock_count': low_stock,
                'in_stock_count': in_stock,
            },
            'product_lines': product_lines_data,
        }

        logger.info(f"Stock report generated: {total_products} products, KES {total_stock_value:,.2f} value")
        return report

    @staticmethod
    def generate_stock_report_for_date(target_date: date) -> Dict:
        """
        Generate stock report as it was on a specific date by reconstructing from movements.

        Args:
            target_date: The date for which to generate the report

        Returns:
            Dictionary with stock summary and product details as of target_date
        """
        logger.info(f"Generating historical stock report for {target_date}")

        # Convert date to datetime at end of day in Nairobi timezone
        target_datetime = timezone.make_aware(
            datetime.combine(target_date, time(23, 59, 59))
        )

        # Get all active products
        products = Product.objects.filter(is_active=True).select_related('product_line').order_by('product_line__name', 'prod_name')

        # Build historical quantities for each product
        product_historical_data = {}
        for product in products:
            # Start with current quantity
            current_quantity = product.quantity

            # Get all movements AFTER the target date (we'll reverse these)
            movements_after = InventoryMovement.objects.filter(
                product=product,
                created_at__gt=target_datetime
            ).order_by('created_at')

            # Reverse the movements to get historical quantity
            historical_quantity = current_quantity
            for movement in movements_after:
                # Subtract the change that happened AFTER our target date
                historical_quantity -= movement.quantity_change

            product_historical_data[product.id] = historical_quantity

        # Calculate totals with historical quantities
        total_products = len(product_historical_data)
        total_stock_value = sum(
            (product_historical_data[p.id] * p.cost_price) for p in products
        )
        total_cost_value = sum(
            (product_historical_data[p.id] * p.cost_price) for p in products
        )

        # Stock status counts (using historical quantities)
        out_of_stock = sum(1 for qty in product_historical_data.values() if qty <= 0)
        low_stock = sum(
            1 for p in products
            if product_historical_data[p.id] > 0 and product_historical_data[p.id] <= p.reorder_level
        )
        in_stock = sum(
            1 for p in products
            if product_historical_data[p.id] > p.reorder_level
        )

        # Group by product line
        product_lines_data = []
        product_lines = ProductLine.objects.all().order_by('name')

        for product_line in product_lines:
            line_products = products.filter(product_line=product_line)
            if line_products.exists():
                line_stock_value = sum(
                    (product_historical_data[p.id] * p.cost_price) for p in line_products
                )
                product_lines_data.append({
                    'product_line_name': product_line.name,
                    'product_count': line_products.count(),
                    'total_quantity': sum(product_historical_data[p.id] for p in line_products),
                    'stock_value': float(line_stock_value),
                    'products': [
                        {
                            'id': p.id,
                            'prod_code': p.prod_code,
                            'prod_name': p.prod_name,
                            'sku': p.sku,
                            'barcode': p.barcode,
                            'quantity': product_historical_data[p.id],
                            'reorder_level': p.reorder_level,
                            'current_price': float(p.current_price),
                            'cost_price': float(p.cost_price),
                            'stock_value': float(product_historical_data[p.id] * p.cost_price),
                            'stock_status': StockReportService._get_stock_status_for_quantity(
                                product_historical_data[p.id], p.reorder_level
                            ),
                        }
                        for p in line_products
                    ]
                })

        # Products without product line
        unassigned = products.filter(product_line__isnull=True)
        if unassigned.exists():
            unassigned_stock_value = sum(
                (product_historical_data[p.id] * p.cost_price) for p in unassigned
            )
            product_lines_data.append({
                'product_line_name': 'Unassigned',
                'product_count': unassigned.count(),
                'total_quantity': sum(product_historical_data[p.id] for p in unassigned),
                'stock_value': float(unassigned_stock_value),
                'products': [
                    {
                        'id': p.id,
                        'prod_code': p.prod_code,
                        'prod_name': p.prod_name,
                        'sku': p.sku,
                        'barcode': p.barcode,
                        'quantity': product_historical_data[p.id],
                        'reorder_level': p.reorder_level,
                        'current_price': float(p.current_price),
                        'cost_price': float(p.cost_price),
                        'stock_value': float(product_historical_data[p.id] * p.cost_price),
                        'stock_status': StockReportService._get_stock_status_for_quantity(
                            product_historical_data[p.id], p.reorder_level
                        ),
                    }
                    for p in unassigned
                ]
            })

        report = {
            'generated_at': timezone.now().isoformat(),
            'report_date': target_date.isoformat(),
            'summary': {
                'total_products': total_products,
                'total_stock_value': float(total_stock_value),
                'total_cost_value': float(total_cost_value),
                'out_of_stock_count': out_of_stock,
                'low_stock_count': low_stock,
                'in_stock_count': in_stock,
            },
            'product_lines': product_lines_data,
        }

        logger.info(f"Historical stock report generated for {target_date}: {total_products} products, KES {total_stock_value:,.2f} value")
        return report

    @staticmethod
    def generate_stock_report_with_adjustments(report_date: date) -> Dict:
        """
        Generate stock report with adjustments for a specific date.

        If a reconciliation exists for the date, includes adjustment columns.
        Otherwise, returns normal stock report with empty adjustments.

        Args:
            report_date: Date for the report

        Returns:
            Dictionary with stock report data and adjustments
        """
        logger.info(f"Generating stock report with adjustments for {report_date}")

        # Get base report data
        if report_date == date.today():
            report_data = StockReportService.generate_stock_report()
        else:
            report_data = StockReportService.generate_stock_report_for_date(report_date)

        # Get reconciliation for the date (if exists)
        reconciliation = DailyStockReconciliation.objects.filter(
            reconciliation_date=report_date
        ).prefetch_related('adjustments__product').first()

        # Build adjustment lookup
        adjustments_by_product = {}
        if reconciliation:
            for adj in reconciliation.adjustments.all():
                adjustments_by_product[adj.product.id] = {
                    'quantity_added': adj.quantity_added,
                    'quantity_deducted': adj.quantity_deducted,
                    'notes': adj.notes,
                    'opening_stock': adj.opening_stock,
                    'closing_stock': adj.closing_stock
                }

        # Add adjustment data to each product in report
        for product_line_data in report_data['product_lines']:
            for product in product_line_data['products']:
                product_id = product['id']
                if product_id in adjustments_by_product:
                    adj = adjustments_by_product[product_id]
                    product['quantity_added'] = adj['quantity_added']
                    product['quantity_deducted'] = adj['quantity_deducted']
                    product['adjustment_notes'] = adj['notes']
                    product['opening_stock'] = adj['opening_stock']
                    product['closing_stock'] = adj['closing_stock']
                else:
                    # No adjustments for this product
                    product['quantity_added'] = 0
                    product['quantity_deducted'] = 0
                    product['adjustment_notes'] = ''
                    product['opening_stock'] = product['quantity']
                    product['closing_stock'] = product['quantity']

        # Add reconciliation metadata
        report_data['has_reconciliation'] = reconciliation is not None
        report_data['reconciliation_status'] = reconciliation.status if reconciliation else None
        report_data['reconciliation_confirmed'] = reconciliation.is_confirmed() if reconciliation else False

        return report_data

    @staticmethod
    def generate_stock_report_xlsx() -> tuple[BytesIO, str]:
        """
        Generate stock report as XLSX file.

        Returns:
            Tuple of (BytesIO object, filename)
        """
        logger.info("Generating stock report XLSX")

        report_data = StockReportService.generate_stock_report()

        # Generate filename with timestamp
        timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
        filename = f"Stock_Report_{timestamp}.xlsx"

        # Create workbook
        wb = Workbook()
        wb.remove(wb.active)

        # Create a single sheet with all products and totals
        StockReportService._create_all_products_sheet(wb, report_data)

        # Save to buffer
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        logger.info(f"Stock report XLSX generated: {filename}")
        return output, filename

    @staticmethod
    def generate_stock_report_xlsx_with_adjustments(report_date: date) -> tuple[BytesIO, str]:
        """
        Generate stock report as XLSX file with adjustments for a specific date.

        Includes Added/Deducted columns showing manual adjustments.

        Args:
            report_date: Date for the report

        Returns:
            Tuple of (BytesIO object, filename)
        """
        logger.info(f"Generating stock report XLSX with adjustments for {report_date}")

        report_data = StockReportService.generate_stock_report_with_adjustments(report_date)

        # Generate filename with date
        filename = f"Stock_Report_{report_date.strftime('%Y%m%d')}.xlsx"

        # Create workbook
        wb = Workbook()
        wb.remove(wb.active)

        # Create products sheet with adjustment columns and totals
        StockReportService._create_all_products_sheet_with_adjustments(wb, report_data)

        # Save to buffer
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        logger.info(f"Stock report XLSX with adjustments generated: {filename}")
        return output, filename

    @staticmethod
    def generate_stock_report_xlsx_for_date(target_date: date) -> tuple[BytesIO, str]:
        """
        Generate historical stock report as XLSX file.

        Args:
            target_date: The date for which to generate the report

        Returns:
            Tuple of (BytesIO object, filename)
        """
        logger.info(f"Generating historical stock report XLSX for {target_date}")

        report_data = StockReportService.generate_stock_report_for_date(target_date)

        # Generate filename with date
        filename = f"Stock_Report_{target_date.strftime('%Y%m%d')}.xlsx"

        # Create workbook
        wb = Workbook()
        wb.remove(wb.active)

        # Create a single sheet with all products and totals
        StockReportService._create_all_products_sheet(wb, report_data)

        # Save to buffer
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        logger.info(f"Historical stock report XLSX generated: {filename}")
        return output, filename

    @staticmethod
    def _create_summary_sheet(wb: Workbook, report_data: Dict):
        """Create summary sheet with overall inventory statistics."""
        ws = wb.create_sheet(title="Summary")

        # Define styles
        title_font = Font(name='Calibri', size=16, bold=True)
        header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4A5568', end_color='4A5568', fill_type='solid')

        # Title
        ws['A1'] = f"Inventory Stock Report"
        ws['A1'].font = title_font
        ws['A2'] = f"Generated: {report_data['generated_at']}"

        # Overall summary
        row = 4
        ws[f'A{row}'] = "Overall Summary"
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].fill = header_fill

        row += 1
        summary_data = [
            ['Total Products', report_data['summary']['total_products']],
            ['Total Stock Value (KES)', f"{report_data['summary']['total_stock_value']:,.2f}"],
            ['Total Cost Value (KES)', f"{report_data['summary']['total_cost_value']:,.2f}"],
            ['In Stock', report_data['summary']['in_stock_count']],
            ['Low Stock', report_data['summary']['low_stock_count']],
            ['Out of Stock', report_data['summary']['out_of_stock_count']],
        ]

        for label, value in summary_data:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            row += 1

        # Product Line breakdown
        row += 2
        ws[f'A{row}'] = "Product Line Breakdown"
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].fill = header_fill

        row += 1
        ws[f'A{row}'] = "Product Line"
        ws[f'B{row}'] = "Products"
        ws[f'C{row}'] = "Total Quantity"
        ws[f'D{row}'] = "Stock Value (KES)"
        for col in ['A', 'B', 'C', 'D']:
            ws[f'{col}{row}'].font = header_font
            ws[f'{col}{row}'].fill = PatternFill(start_color='E2E8F0', end_color='E2E8F0', fill_type='solid')

        row += 1
        for product_line in report_data['product_lines']:
            ws[f'A{row}'] = product_line['product_line_name']
            ws[f'B{row}'] = product_line['product_count']
            ws[f'C{row}'] = product_line['total_quantity']
            ws[f'D{row}'] = f"{product_line['stock_value']:,.2f}"
            row += 1

        # Auto-size columns
        for col in ['A', 'B', 'C', 'D']:
            ws.column_dimensions[col].width = 25

    @staticmethod
    def _create_all_products_sheet(wb: Workbook, report_data: Dict):
        """Create a single sheet with all products from all product lines with totals at the bottom."""
        ws = wb.create_sheet(title="Stock Report")

        # Define styles
        header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4A5568', end_color='4A5568', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        total_font = Font(name='Calibri', size=11, bold=True)
        total_fill = PatternFill(start_color='E2E8F0', end_color='E2E8F0', fill_type='solid')

        # Headers - removed Product Line, Product Code, Barcode, Reorder Level
        headers = ['Product Name', 'SKU', 'Quantity', 'Unit Price', 'Stock Value', 'Status']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Data rows - iterate through all product lines
        row_num = 2
        for product_line_data in report_data['product_lines']:
            for product in product_line_data['products']:
                ws.cell(row=row_num, column=1, value=product['prod_name'])
                ws.cell(row=row_num, column=2, value=product['sku'])
                ws.cell(row=row_num, column=3, value=product['quantity'])

                price_cell = ws.cell(row=row_num, column=4, value=product['cost_price'])
                price_cell.number_format = '#,##0.00'

                value_cell = ws.cell(row=row_num, column=5, value=product['stock_value'])
                value_cell.number_format = '#,##0.00'

                status_cell = ws.cell(row=row_num, column=6, value=product['stock_status'])

                # Color code by status
                if product['stock_status'] == 'OUT_OF_STOCK':
                    status_cell.fill = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')
                elif product['stock_status'] == 'LOW_STOCK':
                    status_cell.fill = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid')
                else:
                    status_cell.fill = PatternFill(start_color='D1FAE5', end_color='D1FAE5', fill_type='solid')

                row_num += 1

        # Add totals row at the bottom
        total_row = row_num
        ws.cell(row=total_row, column=1, value='TOTAL')
        ws.cell(row=total_row, column=1).font = total_font
        ws.cell(row=total_row, column=1).fill = total_fill

        # Calculate total quantity
        total_quantity = report_data['summary']['total_products']
        ws.cell(row=total_row, column=3, value=total_quantity)
        ws.cell(row=total_row, column=3).font = total_font
        ws.cell(row=total_row, column=3).fill = total_fill

        # Total stock value
        total_value_cell = ws.cell(row=total_row, column=5, value=report_data['summary']['total_stock_value'])
        total_value_cell.font = total_font
        total_value_cell.fill = total_fill
        total_value_cell.number_format = '#,##0.00'

        # Auto-size columns
        column_widths = {'A': 40, 'B': 15, 'C': 12, 'D': 15, 'E': 15, 'F': 15}
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

    @staticmethod
    def _create_product_line_sheet(wb: Workbook, product_line_data: Dict):
        """Create detailed sheet for a product line (DEPRECATED - kept for compatibility)."""
        # Sanitize sheet name
        sheet_name = product_line_data['product_line_name'][:30]
        ws = wb.create_sheet(title=sheet_name)

        # Define styles
        header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4A5568', end_color='4A5568', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')

        # Headers
        headers = ['Product Code', 'Product Name', 'SKU', 'Barcode', 'Quantity', 'Reorder Level', 'Unit Price', 'Stock Value', 'Status']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Data rows
        row_num = 2
        for product in product_line_data['products']:
            ws.cell(row=row_num, column=1, value=product['prod_code'])
            ws.cell(row=row_num, column=2, value=product['prod_name'])
            ws.cell(row=row_num, column=3, value=product['sku'])
            ws.cell(row=row_num, column=4, value=product['barcode'] or '')
            ws.cell(row=row_num, column=5, value=product['quantity'])
            ws.cell(row=row_num, column=6, value=product['reorder_level'])

            price_cell = ws.cell(row=row_num, column=7, value=product['cost_price'])
            price_cell.number_format = '#,##0.00'

            value_cell = ws.cell(row=row_num, column=8, value=product['stock_value'])
            value_cell.number_format = '#,##0.00'

            status_cell = ws.cell(row=row_num, column=9, value=product['stock_status'])

            # Color code by status
            if product['stock_status'] == 'OUT_OF_STOCK':
                status_cell.fill = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')
            elif product['stock_status'] == 'LOW_STOCK':
                status_cell.fill = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid')
            else:
                status_cell.fill = PatternFill(start_color='D1FAE5', end_color='D1FAE5', fill_type='solid')

            row_num += 1

        # Auto-size columns
        column_widths = {'A': 15, 'B': 35, 'C': 15, 'D': 18, 'E': 12, 'F': 15, 'G': 15, 'H': 15, 'I': 15}
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

    @staticmethod
    def _get_stock_status(product: Product) -> str:
        """Get stock status label for a product."""
        if product.quantity <= 0:
            return 'OUT_OF_STOCK'
        elif product.quantity <= product.reorder_level:
            return 'LOW_STOCK'
        else:
            return 'IN_STOCK'

    @staticmethod
    def _get_stock_status_for_quantity(quantity: int, reorder_level: int) -> str:
        """Get stock status label for a given quantity and reorder level."""
        if quantity <= 0:
            return 'OUT_OF_STOCK'
        elif quantity <= reorder_level:
            return 'LOW_STOCK'
        else:
            return 'IN_STOCK'

    @staticmethod
    def _create_all_products_sheet_with_adjustments(wb: Workbook, report_data: Dict):
        """
        Create a single sheet with all products including Added/Deducted columns with totals at the bottom.

        Columns: Product Name, SKU, Opening Stock, Added, Deducted, Closing Stock,
                 Unit Price, Stock Value, Status, Notes
        """
        ws = wb.create_sheet(title="Stock Report")

        # Define styles
        header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4A5568', end_color='4A5568', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        total_font = Font(name='Calibri', size=11, bold=True)
        total_fill = PatternFill(start_color='E2E8F0', end_color='E2E8F0', fill_type='solid')

        # Headers with adjustment columns - removed Product Line, Product Code, Barcode, Reorder Level
        headers = [
            'Product Name', 'SKU', 'Opening Stock', 'Added', 'Deducted', 'Closing Stock',
            'Unit Price', 'Stock Value', 'Status', 'Notes'
        ]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Data rows - iterate through all product lines
        row_num = 2
        for product_line_data in report_data['product_lines']:
            for product in product_line_data['products']:
                ws.cell(row=row_num, column=1, value=product['prod_name'])
                ws.cell(row=row_num, column=2, value=product['sku'])

                # Adjustment columns
                ws.cell(row=row_num, column=3, value=product.get('opening_stock', product['quantity']))
                ws.cell(row=row_num, column=4, value=product.get('quantity_added', 0))
                ws.cell(row=row_num, column=5, value=product.get('quantity_deducted', 0))
                ws.cell(row=row_num, column=6, value=product.get('closing_stock', product['quantity']))

                price_cell = ws.cell(row=row_num, column=7, value=product['cost_price'])
                price_cell.number_format = '#,##0.00'

                value_cell = ws.cell(row=row_num, column=8, value=product['stock_value'])
                value_cell.number_format = '#,##0.00'

                status_cell = ws.cell(row=row_num, column=9, value=product['stock_status'])

                # Color code by status
                if product['stock_status'] == 'OUT_OF_STOCK':
                    status_cell.fill = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')
                elif product['stock_status'] == 'LOW_STOCK':
                    status_cell.fill = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid')
                else:
                    status_cell.fill = PatternFill(start_color='D1FAE5', end_color='D1FAE5', fill_type='solid')

                # Adjustment notes
                ws.cell(row=row_num, column=10, value=product.get('adjustment_notes', ''))

                row_num += 1

        # Add totals row at the bottom
        total_row = row_num
        ws.cell(row=total_row, column=1, value='TOTAL')
        ws.cell(row=total_row, column=1).font = total_font
        ws.cell(row=total_row, column=1).fill = total_fill

        # Calculate total products
        total_quantity = report_data['summary']['total_products']
        ws.cell(row=total_row, column=6, value=total_quantity)
        ws.cell(row=total_row, column=6).font = total_font
        ws.cell(row=total_row, column=6).fill = total_fill

        # Total stock value
        total_value_cell = ws.cell(row=total_row, column=8, value=report_data['summary']['total_stock_value'])
        total_value_cell.font = total_font
        total_value_cell.fill = total_fill
        total_value_cell.number_format = '#,##0.00'

        # Auto-size columns
        column_widths = {
            'A': 40, 'B': 15, 'C': 15, 'D': 10, 'E': 12, 'F': 15,
            'G': 15, 'H': 15, 'I': 15, 'J': 30
        }
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

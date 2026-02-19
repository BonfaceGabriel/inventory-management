"""
Unified Report Export Service

Generates a single Excel workbook with four sheets:
  1. All Transactions   – every transaction for the date
  2. Combined Orders    – breakdown of combined-order fulfillment with child allocation
  3. Registration Kits  – kits issued that day (each kit = +200 to shop)
  4. Unfulfilled Orders – historical (from unused.xlsx) merged with system data
"""

import os
from datetime import date, datetime
from decimal import Decimal
from io import BytesIO
from typing import List, Dict
import logging

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from django.utils import timezone
from django.db.models import Q, Sum

from payments.models import Transaction, PaymentGateway, CombinedOrder

logger = logging.getLogger(__name__)

# Path to the historical unfulfilled file (pushed manually; monthly-only, temporary)
UNUSED_XLSX_PATH = os.path.join(os.path.dirname(__file__), '..', '..', 'unused.xlsx')

# ---------------------------------------------------------------------------
# Shared style helpers
# ---------------------------------------------------------------------------

THIN_BORDER = Border(
    left=Side(style='thin', color='CBD5E0'),
    right=Side(style='thin', color='CBD5E0'),
    top=Side(style='thin', color='CBD5E0'),
    bottom=Side(style='thin', color='CBD5E0'),
)

HEADER_FONT = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
HEADER_FILL = PatternFill(start_color='4A5568', end_color='4A5568', fill_type='solid')
HEADER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)

SECTION_FONT = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
SECTION_FILL = PatternFill(start_color='0891B2', end_color='0891B2', fill_type='solid')
SECTION_ALIGN = Alignment(horizontal='center', vertical='center')

SUBTOTAL_FONT = Font(name='Calibri', size=11, bold=True)
SUBTOTAL_FILL = PatternFill(start_color='CFFAFE', end_color='CFFAFE', fill_type='solid')

GRAND_FONT = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
GRAND_FILL = PatternFill(start_color='1F2937', end_color='1F2937', fill_type='solid')

NUM_ALIGN = Alignment(horizontal='right', vertical='center')
LEFT_ALIGN = Alignment(horizontal='left', vertical='center')
CENTER_ALIGN = Alignment(horizontal='center', vertical='center')


def _autosize_columns(ws, min_width: int = 12, max_width: int = 60):
    """Set column widths based on the longest cell value in each column."""
    for col_idx, col in enumerate(ws.columns, start=1):
        max_length = 0
        col_letter = get_column_letter(col_idx)
        for cell in col:
            try:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_length + 3, min_width), max_width)


def _apply_header_row(ws, row_num: int, headers: List[str]):
    """Write a styled column-header row and return the next row number."""
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row_num, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER
    return row_num + 1


def _apply_section_banner(ws, row_num: int, text: str, num_cols: int, fill=None):
    """Write a merged section-banner row and return the next row number."""
    ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=num_cols)
    cell = ws.cell(row=row_num, column=1, value=text)
    cell.font = SECTION_FONT
    cell.fill = fill or SECTION_FILL
    cell.alignment = SECTION_ALIGN
    for col in range(1, num_cols + 1):
        ws.cell(row=row_num, column=col).border = THIN_BORDER
    return row_num + 1


def _get_date_range(report_date: date):
    start = timezone.make_aware(datetime.combine(report_date, datetime.min.time()))
    end = timezone.make_aware(datetime.combine(report_date, datetime.max.time()))
    return start, end


# ---------------------------------------------------------------------------
# Sheet 1 – All Transactions (grouped by gateway)
# ---------------------------------------------------------------------------

NUM_COLS_TXN = 5

def _write_txn_row(ws, row: int, txn) -> Decimal:
    """Write a single transaction data row. Returns the remaining amount."""
    remaining = txn.remaining_amount

    ws.cell(row=row, column=1, value=txn.tx_id or '').alignment = LEFT_ALIGN
    ws.cell(row=row, column=1).border = THIN_BORDER

    cell = ws.cell(row=row, column=2, value=float(txn.amount))
    cell.number_format = '#,##0.00'
    cell.alignment = NUM_ALIGN
    cell.border = THIN_BORDER

    cell = ws.cell(row=row, column=3, value=float(txn.amount_fulfilled))
    cell.number_format = '#,##0.00'
    cell.alignment = NUM_ALIGN
    cell.border = THIN_BORDER

    cell = ws.cell(row=row, column=4, value=float(remaining))
    cell.number_format = '#,##0.00'
    cell.alignment = NUM_ALIGN
    cell.border = THIN_BORDER
    if remaining > 0:
        cell.fill = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')

    ws.cell(row=row, column=5,
            value=txn.timestamp.strftime('%Y-%m-%d %H:%M:%S') if txn.timestamp else '').alignment = CENTER_ALIGN
    ws.cell(row=row, column=5).border = THIN_BORDER

    return remaining


def _write_subtotal_row(ws, row: int, label: str, total, fulfilled, remaining, num_cols: int):
    """Write a styled subtotal row."""
    for col in range(1, num_cols + 1):
        ws.cell(row=row, column=col).fill = SUBTOTAL_FILL
        ws.cell(row=row, column=col).font = SUBTOTAL_FONT
        ws.cell(row=row, column=col).border = THIN_BORDER
    ws.cell(row=row, column=1, value=label).alignment = LEFT_ALIGN
    for col, val in [(2, total), (3, fulfilled), (4, remaining)]:
        cell = ws.cell(row=row, column=col, value=float(val))
        cell.number_format = '#,##0.00'
        cell.alignment = NUM_ALIGN


def _write_grand_total_row(ws, row: int, total, fulfilled, remaining, num_cols: int):
    """Write a styled grand-total row."""
    for col in range(1, num_cols + 1):
        ws.cell(row=row, column=col).fill = GRAND_FILL
        ws.cell(row=row, column=col).font = GRAND_FONT
        ws.cell(row=row, column=col).border = THIN_BORDER
    ws.cell(row=row, column=1, value='GRAND TOTAL').alignment = LEFT_ALIGN
    for col, val in [(2, total), (3, fulfilled), (4, remaining)]:
        cell = ws.cell(row=row, column=col, value=float(val))
        cell.number_format = '#,##0.00'
        cell.alignment = NUM_ALIGN


def _build_all_transactions(ws, report_date: date):
    from collections import defaultdict

    headers = ['Transaction ID', 'Amount (KES)', 'Amount Fulfilled (KES)',
               'Amount Remaining (KES)', 'Timestamp']

    start_dt, end_dt = _get_date_range(report_date)

    # All transactions – exclude internal and combined-order parents, sorted by gateway then time
    transactions = Transaction.objects.exclude(
        Q(sender_name__icontains='7974481') | Q(sender_phone__icontains='7974481') |
        Q(combined_order_parent__isnull=False)
    ).filter(
        timestamp__gte=start_dt,
        timestamp__lte=end_dt
    ).select_related('gateway').order_by('gateway__name', 'timestamp')

    # Group by gateway name while preserving order
    gateway_groups: Dict[str, list] = {}
    for txn in transactions:
        gw_name = txn.gateway.name if txn.gateway else 'Unknown Gateway'
        if gw_name not in gateway_groups:
            gateway_groups[gw_name] = []
        gateway_groups[gw_name].append(txn)

    grand_total = Decimal('0.00')
    grand_fulfilled = Decimal('0.00')
    grand_remaining = Decimal('0.00')
    row = 1

    for gw_name, txns in gateway_groups.items():
        # Gateway section banner
        row = _apply_section_banner(ws, row, gw_name, NUM_COLS_TXN)
        # Column headers
        row = _apply_header_row(ws, row, headers)

        gw_total = Decimal('0.00')
        gw_fulfilled = Decimal('0.00')
        gw_remaining = Decimal('0.00')

        for txn in txns:
            remaining = _write_txn_row(ws, row, txn)
            gw_total += txn.amount
            gw_fulfilled += txn.amount_fulfilled
            gw_remaining += remaining
            row += 1

        # Gateway subtotal
        _write_subtotal_row(ws, row, f'{gw_name} Subtotal',
                            gw_total, gw_fulfilled, gw_remaining, NUM_COLS_TXN)
        grand_total += gw_total
        grand_fulfilled += gw_fulfilled
        grand_remaining += gw_remaining
        row += 2  # blank line between gateways

    # Grand total across all gateways
    _write_grand_total_row(ws, row, grand_total, grand_fulfilled, grand_remaining, NUM_COLS_TXN)
    _autosize_columns(ws)


# ---------------------------------------------------------------------------
# Sheet 2 – Combined Orders Breakdown
# ---------------------------------------------------------------------------

# Gateway types that take priority over Till in the waterfall
_PRIORITY_GATEWAY_TYPES = {
    PaymentGateway.GatewayType.MPESA_PAYBILL,
    PaymentGateway.GatewayType.PDQ,
}
_TILL_GATEWAY_TYPE = PaymentGateway.GatewayType.MPESA_TILL


def _waterfall_allocate(children, today_fulfillment: Decimal) -> List[tuple]:
    """
    Allocate today_fulfillment across child transactions using the same waterfall
    logic as ReconciliationV2Service.calculate_till_sales:

      1. Paybill + PDQ children are filled first (priority pool).
      2. Till children receive whatever spills over.
      3. Any remaining gateway types (Bank, Cash, etc.) absorb the final remainder.

    Within each pool the allocation is proportional to child.amount.

    Returns a list of (child, allocated) tuples in display order
    (priority first, then till, then other).
    """
    priority, till, other = [], [], []

    for child in children:
        gw_type = child.gateway.gateway_type if child.gateway else None
        if gw_type in _PRIORITY_GATEWAY_TYPES:
            priority.append(child)
        elif gw_type == _TILL_GATEWAY_TYPE:
            till.append(child)
        else:
            other.append(child)

    priority_pool = sum(c.amount for c in priority)
    till_pool     = sum(c.amount for c in till)
    other_pool    = sum(c.amount for c in other)

    priority_consumed = min(today_fulfillment, priority_pool)
    till_consumed     = min(today_fulfillment - priority_consumed, till_pool)
    other_consumed    = min(today_fulfillment - priority_consumed - till_consumed, other_pool)

    result = []

    def _split(group, pool, consumed):
        for child in group:
            if pool > 0:
                allocated = (child.amount / pool * consumed).quantize(Decimal('0.01'))
            else:
                allocated = Decimal('0.00')
            result.append((child, allocated))

    _split(priority, priority_pool, priority_consumed)
    _split(till,     till_pool,     till_consumed)
    _split(other,    other_pool,    other_consumed)

    return result


def _build_combined_orders(ws, report_date: date):
    headers = ['Child TX ID', 'Gateway', 'Amount (KES)', 'Amount Fulfilled (KES)']

    row = 1
    start_dt, end_dt = _get_date_range(report_date)

    # Combined orders created or active on this date (all non-cancelled statuses)
    combined_orders = CombinedOrder.objects.filter(
        Q(status__in=[CombinedOrder.Status.PENDING,
                      CombinedOrder.Status.IN_PROGRESS,
                      CombinedOrder.Status.PARTIALLY_FULFILLED,
                      CombinedOrder.Status.FULFILLED]),
        Q(fulfilled_at__gte=start_dt, fulfilled_at__lte=end_dt) |
        Q(created_at__gte=start_dt, created_at__lte=end_dt) |
        Q(updated_at__gte=start_dt, updated_at__lte=end_dt)
    ).prefetch_related('transactions__transaction__gateway').order_by('-updated_at')

    if not combined_orders.exists():
        ws.cell(row=1, column=1, value='No combined orders on this date.')
        ws.cell(row=1, column=1).font = Font(italic=True, color='6B7280')
        _autosize_columns(ws)
        return

    num_cols = len(headers)

    for order in combined_orders:
        # Fulfillment that occurred after combining (same as reconciliation service)
        today_fulfillment = order.amount_fulfilled - order.base_amount_fulfilled

        # Section banner
        banner_text = (
            f"{order.combined_order_id}  |  "
            f"Total: {float(order.total_amount):,.2f}  |  "
            f"Fulfilled: {float(order.amount_fulfilled):,.2f}  |  "
            f"Status: {order.get_status_display()}"
        )
        row = _apply_section_banner(ws, row, banner_text, num_cols)
        row = _apply_header_row(ws, row, headers)

        children = [cot.transaction for cot in order.transactions.all()]
        allocations = _waterfall_allocate(children, today_fulfillment)
        order_subtotal_fulfilled = Decimal('0.00')

        for child, allocated in allocations:
            ws.cell(row=row, column=1, value=child.tx_id or '').alignment = LEFT_ALIGN
            ws.cell(row=row, column=1).border = THIN_BORDER

            ws.cell(row=row, column=2,
                    value=child.gateway.name if child.gateway else '').alignment = LEFT_ALIGN
            ws.cell(row=row, column=2).border = THIN_BORDER

            cell = ws.cell(row=row, column=3, value=float(child.amount))
            cell.number_format = '#,##0.00'
            cell.alignment = NUM_ALIGN
            cell.border = THIN_BORDER

            cell = ws.cell(row=row, column=4, value=float(allocated))
            cell.number_format = '#,##0.00'
            cell.alignment = NUM_ALIGN
            cell.border = THIN_BORDER

            order_subtotal_fulfilled += allocated
            row += 1

        # Subtotal row
        for col in range(1, num_cols + 1):
            ws.cell(row=row, column=col).fill = SUBTOTAL_FILL
            ws.cell(row=row, column=col).font = SUBTOTAL_FONT
            ws.cell(row=row, column=col).border = THIN_BORDER
        ws.cell(row=row, column=1, value='Subtotal').alignment = LEFT_ALIGN
        cell = ws.cell(row=row, column=3, value=float(order.total_amount))
        cell.number_format = '#,##0.00'
        cell.alignment = NUM_ALIGN
        cell = ws.cell(row=row, column=4, value=float(order_subtotal_fulfilled))
        cell.number_format = '#,##0.00'
        cell.alignment = NUM_ALIGN

        row += 2  # blank row between orders

    _autosize_columns(ws)


# ---------------------------------------------------------------------------
# Sheet 3 – Registration Kits
# ---------------------------------------------------------------------------

def _build_registration_kits(ws, report_date: date):
    headers = ['Transaction ID', 'Kits Issued', 'Value (KES)', 'Gateway Name']

    KIT_VALUE = Decimal('200.00')  # Same as REGISTRATION_KIT_VALUE in reconciliation_v2

    row = _apply_header_row(ws, 1, headers)
    ws.freeze_panes = 'A2'

    start_dt, end_dt = _get_date_range(report_date)

    # All registration transactions with kits issued on this date
    # (singles + combined-order parents, same logic as reconciliation_v2.calculate_kits)
    base_exclude = Q(sender_name__icontains='7974481') | Q(sender_phone__icontains='7974481')

    reg_txns = Transaction.objects.exclude(
        base_exclude | Q(status=Transaction.OrderStatus.COMBINED_FULFILLED)
    ).filter(
        is_registration=True,
        registration_kit_issued=True,
    ).filter(
        Q(completed_at__gte=start_dt, completed_at__lte=end_dt) |
        Q(timestamp__gte=start_dt, timestamp__lte=end_dt)
    ).select_related('gateway').order_by('timestamp')

    total_kits = 0
    total_value = Decimal('0.00')

    for txn in reg_txns:
        qty = txn.registration_kit_quantity or 0
        value = KIT_VALUE * qty

        ws.cell(row=row, column=1, value=txn.tx_id or '').alignment = LEFT_ALIGN
        ws.cell(row=row, column=1).border = THIN_BORDER

        ws.cell(row=row, column=2, value=qty).alignment = CENTER_ALIGN
        ws.cell(row=row, column=2).border = THIN_BORDER

        cell = ws.cell(row=row, column=3, value=float(value))
        cell.number_format = '#,##0.00'
        cell.alignment = NUM_ALIGN
        cell.border = THIN_BORDER

        ws.cell(row=row, column=4, value=txn.gateway.name if txn.gateway else '').alignment = LEFT_ALIGN
        ws.cell(row=row, column=4).border = THIN_BORDER

        total_kits += qty
        total_value += value
        row += 1

    if total_kits == 0:
        ws.cell(row=row, column=1, value='No registration kits issued on this date.')
        ws.cell(row=row, column=1).font = Font(italic=True, color='6B7280')
        _autosize_columns(ws)
        return

    # Summary row
    row += 1
    for col in range(1, 5):
        ws.cell(row=row, column=col).fill = GRAND_FILL
        ws.cell(row=row, column=col).font = GRAND_FONT
        ws.cell(row=row, column=col).border = THIN_BORDER

    ws.cell(row=row, column=1, value='TOTAL').alignment = LEFT_ALIGN
    ws.cell(row=row, column=2, value=total_kits).alignment = CENTER_ALIGN
    cell = ws.cell(row=row, column=3, value=float(total_value))
    cell.number_format = '#,##0.00'
    cell.alignment = NUM_ALIGN
    ws.cell(row=row, column=4, value=f'{total_kits} kit(s) × 200 KES').alignment = LEFT_ALIGN
    _autosize_columns(ws)


# ---------------------------------------------------------------------------
# Sheet 4 – Unfulfilled Orders
# ---------------------------------------------------------------------------

def _read_unused_xlsx() -> Dict[str, Decimal]:
    """
    Read the historical unfulfilled file (unused.xlsx).
    Returns {tx_id: amount}.  Returns empty dict if file missing.
    """
    path = os.path.abspath(UNUSED_XLSX_PATH)
    if not os.path.exists(path):
        logger.info("unused.xlsx not found at %s — skipping historical section", path)
        return {}

    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        result = {}
        for row in ws.iter_rows(min_row=2, values_only=True):  # skip header
            if row[0] is None:
                continue
            tx_id = str(row[0]).strip()
            amount = Decimal(str(row[1])) if row[1] is not None else Decimal('0.00')
            if tx_id:
                result[tx_id] = amount
        wb.close()
        return result
    except Exception as e:
        logger.error("Failed to read unused.xlsx: %s", e)
        return {}


def _build_unfulfilled_orders(ws, report_date: date):
    headers = ['Transaction ID', 'Amount (KES)', 'Status', 'Source']

    unfulfilled_statuses = [
        Transaction.OrderStatus.NOT_PROCESSED,
        Transaction.OrderStatus.PROCESSING,
    ]

    # --------------- Section 1: Historical (from unused.xlsx) ---------------
    row = _apply_section_banner(ws, 1, 'HISTORICAL UNUSED', 4)
    row = _apply_header_row(ws, row, headers)

    historical_data = _read_unused_xlsx()
    hist_total = Decimal('0.00')
    hist_count = 0

    if historical_data:
        # Batch-check which of those TX IDs are still unfulfilled
        tx_ids = list(historical_data.keys())
        still_unfulfilled = set(
            Transaction.objects.filter(
                tx_id__in=tx_ids,
                status__in=unfulfilled_statuses
            ).values_list('tx_id', flat=True)
        )

        for tx_id in tx_ids:
            if tx_id not in still_unfulfilled:
                continue  # already fulfilled or cancelled – skip

            amount = historical_data[tx_id]
            ws.cell(row=row, column=1, value=tx_id).alignment = LEFT_ALIGN
            ws.cell(row=row, column=1).border = THIN_BORDER

            cell = ws.cell(row=row, column=2, value=float(amount))
            cell.number_format = '#,##0.00'
            cell.alignment = NUM_ALIGN
            cell.border = THIN_BORDER

            # Pull live status
            try:
                live_txn = Transaction.objects.get(tx_id=tx_id)
                ws.cell(row=row, column=3, value=live_txn.get_status_display()).alignment = CENTER_ALIGN
            except Transaction.DoesNotExist:
                ws.cell(row=row, column=3, value='Unknown').alignment = CENTER_ALIGN
            ws.cell(row=row, column=3).border = THIN_BORDER

            ws.cell(row=row, column=4, value='Uploaded File').alignment = CENTER_ALIGN
            ws.cell(row=row, column=4).border = THIN_BORDER
            ws.cell(row=row, column=4).fill = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')

            hist_total += amount
            hist_count += 1
            row += 1
    else:
        ws.cell(row=row, column=1, value='No historical file available.').font = Font(italic=True, color='6B7280')
        row += 1

    # Historical subtotal
    for col in range(1, 5):
        ws.cell(row=row, column=col).fill = SUBTOTAL_FILL
        ws.cell(row=row, column=col).font = SUBTOTAL_FONT
        ws.cell(row=row, column=col).border = THIN_BORDER
    ws.cell(row=row, column=1, value=f'Historical Subtotal ({hist_count})').alignment = LEFT_ALIGN
    cell = ws.cell(row=row, column=2, value=float(hist_total))
    cell.number_format = '#,##0.00'
    cell.alignment = NUM_ALIGN
    row += 2  # blank

    # --------------- Section 2: Today onwards (from system) -----------------
    row = _apply_section_banner(ws, row, 'UNUSED TODAY', 4)
    row = _apply_header_row(ws, row, headers)

    start_dt, _ = _get_date_range(report_date)

    # Paybill transactions from report_date onwards that are still unfulfilled
    # (only paybill – till transactions don't carry over as "unused")
    system_unfulfilled = Transaction.objects.exclude(
        Q(sender_name__icontains='7974481') | Q(sender_phone__icontains='7974481') |
        Q(combined_order_parent__isnull=False) |
        Q(status=Transaction.OrderStatus.COMBINED_FULFILLED)
    ).filter(
        timestamp__gte=start_dt,
        status__in=unfulfilled_statuses,
        gateway__gateway_type=PaymentGateway.GatewayType.MPESA_PAYBILL
    ).select_related('gateway').order_by('timestamp')

    sys_total = Decimal('0.00')
    sys_count = 0

    for txn in system_unfulfilled:
        ws.cell(row=row, column=1, value=txn.tx_id or '').alignment = LEFT_ALIGN
        ws.cell(row=row, column=1).border = THIN_BORDER

        cell = ws.cell(row=row, column=2, value=float(txn.amount))
        cell.number_format = '#,##0.00'
        cell.alignment = NUM_ALIGN
        cell.border = THIN_BORDER

        ws.cell(row=row, column=3, value=txn.get_status_display()).alignment = CENTER_ALIGN
        ws.cell(row=row, column=3).border = THIN_BORDER

        ws.cell(row=row, column=4, value='System').alignment = CENTER_ALIGN
        ws.cell(row=row, column=4).border = THIN_BORDER
        ws.cell(row=row, column=4).fill = PatternFill(start_color='DBEAFE', end_color='DBEAFE', fill_type='solid')

        sys_total += txn.amount
        sys_count += 1
        row += 1

    if sys_count == 0:
        ws.cell(row=row, column=1, value='No unfulfilled paybill transactions from this date.').font = Font(italic=True, color='6B7280')
        row += 1

    # System subtotal
    for col in range(1, 5):
        ws.cell(row=row, column=col).fill = SUBTOTAL_FILL
        ws.cell(row=row, column=col).font = SUBTOTAL_FONT
        ws.cell(row=row, column=col).border = THIN_BORDER
    ws.cell(row=row, column=1, value=f'System Subtotal ({sys_count})').alignment = LEFT_ALIGN
    cell = ws.cell(row=row, column=2, value=float(sys_total))
    cell.number_format = '#,##0.00'
    cell.alignment = NUM_ALIGN
    row += 2

    # Grand total
    for col in range(1, 5):
        ws.cell(row=row, column=col).fill = GRAND_FILL
        ws.cell(row=row, column=col).font = GRAND_FONT
        ws.cell(row=row, column=col).border = THIN_BORDER
    ws.cell(row=row, column=1, value='GRAND TOTAL').alignment = LEFT_ALIGN
    cell = ws.cell(row=row, column=2, value=float(hist_total + sys_total))
    cell.number_format = '#,##0.00'
    cell.alignment = NUM_ALIGN
    _autosize_columns(ws)


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------

class TransactionExportService:
    """Generates the unified daily report workbook."""

    @staticmethod
    def generate_unified_report(report_date: date) -> BytesIO:
        """
        Build a single Excel workbook with all four report sheets.

        Args:
            report_date: The date to generate the report for.

        Returns:
            BytesIO containing the .xlsx file.
        """
        logger.info("Generating unified report for %s", report_date)

        wb = Workbook()

        # Sheet 1 – All Transactions
        ws1 = wb.active
        ws1.title = "All Transactions"
        _build_all_transactions(ws1, report_date)

        # Sheet 2 – Combined Orders
        ws2 = wb.create_sheet("Combined Orders")
        _build_combined_orders(ws2, report_date)

        # Sheet 3 – Registration Kits
        ws3 = wb.create_sheet("Registration Kits")
        _build_registration_kits(ws3, report_date)

        # Sheet 4 – Unfulfilled Orders
        ws4 = wb.create_sheet("Unfulfilled Orders")
        _build_unfulfilled_orders(ws4, report_date)

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        logger.info("Unified report for %s generated successfully", report_date)
        return output

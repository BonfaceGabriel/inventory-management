from io import BytesIO
from openpyxl import load_workbook
from django.test import TestCase
from utils.xlsx_generator import XlsxGenerator


class TestXlsxGenerator(TestCase):
    def test_from_data_returns_bytesio(self):
        data = [{'name': 'Alice', 'score': 95}, {'name': 'Bob', 'score': 87}]
        columns = [
            {'key': 'name', 'header': 'Name', 'width': 20},
            {'key': 'score', 'header': 'Score', 'width': 10, 'align': 'right'},
        ]
        buf = XlsxGenerator.from_data(data, columns, sheet_name='Scores', title='Test Report')
        self.assertIsInstance(buf, BytesIO)
        self.assertGreater(len(buf.getvalue()), 100)
        wb = load_workbook(buf)
        ws = wb.active
        self.assertEqual(ws.title, 'Scores')
        self.assertEqual(ws.cell(1, 1).value, 'Test Report')
        self.assertEqual(ws.cell(2, 1).value, 'Name')
        self.assertEqual(ws.cell(3, 1).value, 'Alice')
        self.assertEqual(ws.cell(4, 2).value, 87)

    def test_from_data_empty(self):
        buf = XlsxGenerator.from_data([], [], sheet_name='Empty')
        self.assertIsInstance(buf, BytesIO)
        wb = load_workbook(buf)
        ws = wb.active
        self.assertEqual(ws.title, 'Empty')

    def test_multi_sheet_returns_bytesio(self):
        sheets = [
            {
                'name': 'Summary',
                'data': [{'key': 'X', 'val': 100}],
                'columns': [{'key': 'key', 'header': 'Key', 'width': 10}, {'key': 'val', 'header': 'Value', 'width': 10}],
                'title': 'Summary Sheet',
            },
            {
                'name': 'Details',
                'data': [{'item': 'A', 'qty': 5}],
                'columns': [{'key': 'item', 'header': 'Item', 'width': 15}, {'key': 'qty', 'header': 'Qty', 'width': 10}],
                'title': 'Details Sheet',
            },
        ]
        buf = XlsxGenerator.multi_sheet(sheets)
        self.assertIsInstance(buf, BytesIO)
        wb = load_workbook(buf)
        self.assertEqual(len(wb.sheetnames), 2)
        self.assertIn('Summary', wb.sheetnames)
        self.assertIn('Details', wb.sheetnames)

    def test_multi_sheet_empty(self):
        buf = XlsxGenerator.multi_sheet([])
        self.assertIsInstance(buf, BytesIO)

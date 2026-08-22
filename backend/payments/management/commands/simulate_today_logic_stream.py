from __future__ import annotations

import hashlib
import time
from decimal import Decimal
from io import BytesIO

from django.core.management import call_command
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction as db_transaction
from django.utils import timezone
from openpyxl import load_workbook

from payments.models import (
    CombinedOrder,
    DailyStockReconciliation,
    EndOfDayValueReconciliation,
    MerchandiseCatalogItem,
    MerchandiseOrder,
    PaymentGateway,
    Product,
    StockAdjustmentItem,
    Transaction,
    User,
)
from payments.services.admin_service import AdminService
from payments.services.combined_order_service import CombinedOrderService
from payments.services.eod_value_reconciliation_service import EndOfDayValueReconciliationService
from payments.services.fulfillment_service import FulfillmentService
from payments.services.merchandise_service import MerchandiseService
from payments.services.reconciliation_workflow_service import ReconciliationWorkflowService
from payments.services.registration_kit_service import RegistrationKitService
from payments.services.stock_report_service import StockReportService
from payments.services.stock_take_service import StockTakeService


class Command(BaseCommand):
    help = (
        "Simulate today's core logic streams (forward + reverse flows) and verify "
        "EOD X values match stock report X components."
    )

    def add_arguments(self, parser):
        parser.add_argument(
            "--clear",
            action="store_true",
            help="Delete prior simulation rows (SIMTODAY-*) before generating new ones.",
        )

    def handle(self, *args, **options):
        today = timezone.localdate()
        active_session = StockTakeService.get_active_session()
        if active_session:
            raise CommandError(
                f"Active stock-take session {active_session.session_id} detected. "
                "Complete/cancel it first, then rerun this command."
            )

        if options["clear"]:
            self._clear_previous_simulation(today)

        admin, processor, issuer = self._ensure_users()
        gateways = self._ensure_gateways()
        products = self._ensure_products()

        self.stdout.write(self.style.WARNING(f"Running today's simulation for {today}..."))

        scenario = self._run_streams(
            admin=admin,
            processor=processor,
            issuer=issuer,
            gateways=gateways,
            products=products,
        )

        eod_record, expected_x_components = self._sync_and_validate_eod(admin, today)
        self._validate_stock_report_eod_block(today, eod_record, expected_x_components)

        self.stdout.write(self.style.SUCCESS("\nSimulation complete."))
        self.stdout.write(self.style.SUCCESS("Actions executed:"))
        for line in scenario:
            self.stdout.write(f"  - {line}")
        self.stdout.write(
            self.style.SUCCESS(
                "EOD verification: PASS (X components in EOD and stock report are aligned)."
            )
        )

    def _clear_previous_simulation(self, today):
        self.stdout.write(self.style.WARNING("Clearing previous SIMTODAY data..."))
        MerchandiseOrder.objects.filter(transaction__tx_id__startswith="SIMTODAY-").delete()
        Transaction.objects.filter(tx_id__startswith="SIMTODAY-").delete()
        CombinedOrder.objects.filter(created_by="sim_today").delete()
        StockTakeSessionQuery = DailyStockReconciliation.objects.filter(reconciliation_date=today)
        if StockTakeSessionQuery.exists():
            # Keep reconciliation rows if already used elsewhere by user, only clean EOD draft.
            pass
        EndOfDayValueReconciliation.objects.filter(
            reconciliation_date=today,
            status=EndOfDayValueReconciliation.Status.DRAFT,
        ).delete()
        self.stdout.write(self.style.WARNING("Previous simulation data cleared."))

    def _ensure_users(self):
        admin, _ = User.objects.get_or_create(
            username="sim_today_admin",
            defaults={"role": User.Role.ADMIN, "password": "x"},
        )
        if admin.role != User.Role.ADMIN:
            admin.role = User.Role.ADMIN
            admin.save(update_fields=["role"])

        processor, _ = User.objects.get_or_create(
            username="sim_today_processor",
            defaults={"role": User.Role.PROCESSOR, "password": "x"},
        )
        if processor.role != User.Role.PROCESSOR:
            processor.role = User.Role.PROCESSOR
            processor.save(update_fields=["role"])

        issuer, _ = User.objects.get_or_create(
            username="sim_today_issuer",
            defaults={"role": User.Role.ISSUER, "password": "x"},
        )
        if issuer.role != User.Role.ISSUER:
            issuer.role = User.Role.ISSUER
            issuer.save(update_fields=["role"])

        return admin, processor, issuer

    def _ensure_gateways(self):
        till_products, _ = PaymentGateway.objects.get_or_create(
            name="Till Products",
            defaults={
                "gateway_type": PaymentGateway.GatewayType.MPESA_TILL,
                "gateway_number": "111111",
                "settlement_type": PaymentGateway.SettlementType.NONE,
                "is_active": True,
            },
        )
        merch_gateway, _ = PaymentGateway.objects.get_or_create(
            name="Till Merchandise",
            defaults={
                "gateway_type": PaymentGateway.GatewayType.MERCHANDISE,
                "gateway_number": "222222",
                "settlement_type": PaymentGateway.SettlementType.NONE,
                "is_active": True,
            },
        )
        paybill, _ = PaymentGateway.objects.get_or_create(
            name="Paybill",
            defaults={
                "gateway_type": PaymentGateway.GatewayType.MPESA_PAYBILL,
                "gateway_number": "333333",
                "settlement_type": PaymentGateway.SettlementType.NONE,
                "is_active": True,
            },
        )
        pdq, _ = PaymentGateway.objects.get_or_create(
            name="PDQ",
            defaults={
                "gateway_type": PaymentGateway.GatewayType.PDQ,
                "gateway_number": "444444",
                "settlement_type": PaymentGateway.SettlementType.NONE,
                "is_active": True,
            },
        )
        return {
            "till": till_products,
            "merch": merch_gateway,
            "paybill": paybill,
            "pdq": pdq,
        }

    def _ensure_products(self):
        product_a, _ = Product.objects.get_or_create(
            prod_code="SIMP001",
            defaults={
                "prod_name": "Simulation Product A",
                "sku": "SIMP001",
                "sku_name": "Unit",
                "barcode": "SIMP001",
                "current_price": Decimal("1000.00"),
                "cost_price": Decimal("700.00"),
                "current_pv": Decimal("10.00"),
                "quantity": 200,
                "is_active": True,
            },
        )
        product_b, _ = Product.objects.get_or_create(
            prod_code="SIMP002",
            defaults={
                "prod_name": "Simulation Product B",
                "sku": "SIMP002",
                "sku_name": "Unit",
                "barcode": "SIMP002",
                "current_price": Decimal("500.00"),
                "cost_price": Decimal("300.00"),
                "current_pv": Decimal("5.00"),
                "quantity": 300,
                "is_active": True,
            },
        )
        reg_kit, _ = Product.objects.get_or_create(
            prod_code="REG_KIT_001",
            defaults={
                "prod_name": "Registration Kit",
                "sku": "REG_KIT_001",
                "sku_name": "Kit",
                "barcode": "REG_KIT_001",
                "current_price": Decimal("2900.00"),
                "cost_price": Decimal("2700.00"),
                "current_pv": Decimal("0.00"),
                "quantity": 50,
                "is_active": True,
            },
        )
        return {"a": product_a, "b": product_b, "kit": reg_kit}

    def _make_tx(self, tx_id: str, amount: Decimal, gateway: PaymentGateway) -> Transaction:
        unique_hash = hashlib.sha256(
            f"{tx_id}|{amount}|{timezone.now().isoformat()}".encode()
        ).hexdigest()
        return Transaction.objects.create(
            tx_id=tx_id,
            amount=amount,
            sender_name="Simulation Customer",
            sender_phone="0712345678",
            timestamp=timezone.now(),
            gateway=gateway,
            gateway_type=gateway.gateway_type,
            destination_number=gateway.gateway_number,
            status=Transaction.OrderStatus.NOT_PROCESSED,
            confidence=Decimal("0.99"),
            unique_hash=unique_hash,
            notes="SIMTODAY generated",
        )

    def _run_streams(self, admin, processor, issuer, gateways, products):
        actions = []
        with db_transaction.atomic():
            # 1) Raw not-processed transaction
            tx_not_processed = self._make_tx("SIMTODAY-NP-001", Decimal("2500.00"), gateways["paybill"])
            actions.append(f"Created NOT_PROCESSED transaction {tx_not_processed.tx_id}")

            # Gateway diversity: explicit processing/cancelled/manual-like transactions
            tx_pdq = self._make_tx("SIMTODAY-PDQ-001", Decimal("1800.00"), gateways["pdq"])
            tx_pdq.status = Transaction.OrderStatus.PROCESSING
            tx_pdq.save(update_fields=["status"])
            actions.append(f"Created PROCESSING PDQ transaction {tx_pdq.tx_id}")

            # 2) Issuance cancel path (no inventory deduction)
            tx_cancel = self._make_tx("SIMTODAY-CAN-001", Decimal("3000.00"), gateways["till"])
            FulfillmentService.activate_issuance(tx_cancel.id, activated_by_user=processor)
            FulfillmentService.scan_barcode(
                tx_cancel.id, {"prod_code": products["a"].prod_code, "quantity": 1}, scanned_by_user=processor
            )
            FulfillmentService.cancel_issuance(tx_cancel.id, reason="simulate cancel before complete")
            actions.append(f"Ran cancel-issuance flow on {tx_cancel.tx_id}")

            # 3) Full fulfillment path
            tx_full = self._make_tx("SIMTODAY-FUL-001", Decimal("2000.00"), gateways["till"])
            FulfillmentService.activate_issuance(tx_full.id, activated_by_user=processor)
            FulfillmentService.scan_barcode(
                tx_full.id, {"prod_code": products["a"].prod_code, "quantity": 2}, scanned_by_user=processor
            )
            FulfillmentService.complete_issuance(tx_full.id, completed_by_user=processor)
            actions.append(f"Completed fulfillment flow on {tx_full.tx_id}")

            # 4) Partial -> resume -> complete
            tx_partial = self._make_tx("SIMTODAY-PAR-001", Decimal("3000.00"), gateways["paybill"])
            FulfillmentService.activate_issuance(tx_partial.id, activated_by_user=processor)
            FulfillmentService.scan_barcode(
                tx_partial.id, {"prod_code": products["a"].prod_code, "quantity": 1}, scanned_by_user=processor
            )
            FulfillmentService.complete_issuance(tx_partial.id, completed_by_user=processor)
            FulfillmentService.activate_issuance(tx_partial.id, activated_by_user=processor)
            FulfillmentService.scan_barcode(
                tx_partial.id, {"prod_code": products["b"].prod_code, "quantity": 4}, scanned_by_user=processor
            )
            FulfillmentService.complete_issuance(tx_partial.id, completed_by_user=processor)
            actions.append(f"Ran partial->resume->complete flow on {tx_partial.tx_id}")

            # 5) Admin reversal on fulfilled transaction
            tx_reverse = self._make_tx("SIMTODAY-REV-001", Decimal("1000.00"), gateways["till"])
            FulfillmentService.activate_issuance(tx_reverse.id, activated_by_user=processor)
            FulfillmentService.scan_barcode(
                tx_reverse.id, {"prod_code": products["a"].prod_code, "quantity": 1}, scanned_by_user=processor
            )
            FulfillmentService.complete_issuance(tx_reverse.id, completed_by_user=processor)
            AdminService.cancel_fulfilled_transaction(
                transaction_id=tx_reverse.id,
                cancelled_by_user=admin,
                reason="simulate reverse movement",
            )
            actions.append(f"Ran admin cancel reversal flow on {tx_reverse.tx_id}")

            # 6) Registration-kit issuance path
            tx_reg = self._make_tx("SIMTODAY-REG-001", Decimal("5000.00"), gateways["till"])
            tx_reg.is_registration = True
            tx_reg.save(update_fields=["is_registration"])
            RegistrationKitService.issue_registration_kit(tx_reg.id, quantity=1, issued_by=processor.username)
            actions.append(f"Issued registration kit for {tx_reg.tx_id}")

            # 7) Combined-order create + complete + revert
            tx_co_1 = self._make_tx("SIMTODAY-CO-001", Decimal("2000.00"), gateways["till"])
            tx_co_2 = self._make_tx("SIMTODAY-CO-002", Decimal("1500.00"), gateways["paybill"])
            FulfillmentService.activate_issuance(tx_co_1.id, activated_by_user=processor)
            FulfillmentService.scan_barcode(
                tx_co_1.id, {"prod_code": products["a"].prod_code, "quantity": 1}, scanned_by_user=processor
            )
            FulfillmentService.complete_issuance(tx_co_1.id, completed_by_user=processor)
            combined = CombinedOrderService.create_combined_order(
                transaction_ids=[tx_co_1.id, tx_co_2.id],
                created_by="sim_today",
                created_by_user=processor,
            )
            co_id = combined["combined_order_id"]
            CombinedOrderService.activate_combined_order(co_id, "sim_today")
            CombinedOrderService.scan_product_to_combined_order_staged(co_id, products["b"].id, 2, "sim_today")
            CombinedOrderService.complete_combined_order(co_id, "sim_today")
            CombinedOrderService.revert_combined_order(co_id, "sim_today", "simulation revert check")
            actions.append(f"Ran combined-order complete+revert on {co_id}")

            # 8) Merchandise fulfillment path (if merchandise catalog present)
            merch_action = self._run_merchandise_stream(processor, gateways["merch"])
            actions.append(merch_action)

            # 9) Stock take item-based + kit-only session
            session = StockTakeService.create_session(created_by=issuer.username, notes="SIMTODAY item-based")
            StockTakeService.scan_product(session.session_id, products["b"].id, 3, issuer.username)
            StockTakeService.complete_session(session.session_id, issuer.username)
            # Session IDs are second-based; ensure next one has a unique timestamp suffix.
            time.sleep(1.1)
            kit_session = StockTakeService.create_session(created_by=issuer.username, notes="SIMTODAY kit-only")
            StockTakeService.update_kit_quantity(kit_session.session_id, 2)
            StockTakeService.complete_session(kit_session.session_id, issuer.username)
            actions.append("Completed stock-take item-based and kit-only sessions")

        return actions

    def _run_merchandise_stream(self, processor_user, merch_gateway):
        if not MerchandiseCatalogItem.objects.filter(is_active=True).exists():
            try:
                call_command("seed_merchandise_catalog")
            except Exception:
                return "Skipped merchandise stream (catalog unavailable)"

        tx_merch = self._make_tx("SIMTODAY-MERCH-001", Decimal("3500.00"), merch_gateway)
        order = MerchandiseService.create_pending_order_for_transaction(tx_merch)
        if not order:
            return "Skipped merchandise stream (pending order not created)"

        stock_rows = [r for r in MerchandiseService.get_stock_rows() if r["quantity"] > 0]
        if not stock_rows:
            # Ensure at least one stock variant is available
            first_item = MerchandiseCatalogItem.objects.filter(is_active=True).first()
            if not first_item:
                return "Skipped merchandise stream (no active items)"
            adjustment = {"item_code": first_item.code, "quantity_change": 5}
            if first_item.item_type == MerchandiseCatalogItem.ItemType.TSHIRT:
                colors = list(first_item.options.filter(option_type="COLOR").values_list("value", flat=True))
                sizes = list(first_item.options.filter(option_type="SIZE").values_list("value", flat=True))
                if colors:
                    adjustment["color"] = colors[0]
                if sizes:
                    adjustment["size"] = sizes[0]
            elif first_item.item_type == MerchandiseCatalogItem.ItemType.HAT:
                colors = list(first_item.options.filter(option_type="COLOR").values_list("value", flat=True))
                if colors:
                    adjustment["color"] = colors[0]
            MerchandiseService.adjust_stock([adjustment], processor_user, notes="SIMTODAY ensure stock")
            stock_rows = [r for r in MerchandiseService.get_stock_rows() if r["quantity"] > 0]
            if not stock_rows:
                return "Skipped merchandise stream (could not prepare stock)"

        row = stock_rows[0]
        payload = {"item_code": row["item_code"], "quantity": 1}
        if row.get("color"):
            payload["color"] = row["color"]
        if row.get("size"):
            payload["size"] = row["size"]
        MerchandiseService.fulfill_order(order, [payload], processor_user)
        return f"Completed merchandise fulfillment flow on {tx_merch.tx_id}"

    def _sync_and_validate_eod(self, admin_user, today):
        reconciliation = ReconciliationWorkflowService.get_or_create_reconciliation(today, admin_user)
        # Touch at least one manual adjustment to exercise adjustment stream.
        first_adjustment = reconciliation.adjustments.select_related("product").first()
        if first_adjustment:
            ReconciliationWorkflowService.update_adjustment(
                reconciliation_id=str(reconciliation.id),
                product_id=first_adjustment.product_id,
                quantity_added=0,
                quantity_deducted=0,
                notes="SIMTODAY refresh",
            )

        eod = EndOfDayValueReconciliationService.get_or_create_today(admin_user)
        eod = EndOfDayValueReconciliationService.update_today_inputs(
            admin_user,
            {
                "stock_value": Decimal("0.00"),
                "bk_stock": Decimal("0.00"),
                "duplicated": Decimal("0.00"),
                "hq_value": Decimal("0.00"),
                "kitengela_value": Decimal("0.00"),
                "kitui_value": Decimal("0.00"),
                "nakuru_value": Decimal("0.00"),
            },
        )

        adjustments = reconciliation.adjustments.select_related("product").all()
        opening_total = sum((adj.opening_stock * adj.product.cost_price) for adj in adjustments)
        replenished_total = sum((adj.quantity_replenished * adj.product.cost_price) for adj in adjustments)
        sales_total = sum((Decimal(adj.sales or 0) * adj.product.cost_price) for adj in adjustments)
        expected_x = opening_total + replenished_total - sales_total

        if eod.opening_stock_value != opening_total:
            raise CommandError(
                f"EOD opening_stock_value mismatch: expected {opening_total}, got {eod.opening_stock_value}"
            )
        if eod.replenished_value != replenished_total:
            raise CommandError(
                f"EOD replenished_value mismatch: expected {replenished_total}, got {eod.replenished_value}"
            )
        if eod.sales_value != sales_total:
            raise CommandError(
                f"EOD sales_value mismatch: expected {sales_total}, got {eod.sales_value}"
            )
        if eod.x_value != expected_x:
            raise CommandError(f"EOD x_value mismatch: expected {expected_x}, got {eod.x_value}")

        return eod, {
            "opening": opening_total,
            "replenished": replenished_total,
            "sales": sales_total,
            "x": expected_x,
        }

    def _validate_stock_report_eod_block(self, today, eod, expected):
        xlsx_buffer, _ = StockReportService.generate_stock_report_xlsx_with_adjustments(today)
        wb = load_workbook(filename=BytesIO(xlsx_buffer.getvalue()))
        ws = wb["Stock Report"]

        label_to_value = {}
        for row in ws.iter_rows(min_row=1, max_col=2, values_only=True):
            label, value = row
            if isinstance(label, str):
                label_to_value[label] = value

        required = {
            "Opening Stock Value": float(expected["opening"]),
            "Replenished Value": float(expected["replenished"]),
            "Sales Value": float(expected["sales"]),
            "X Value": float(expected["x"]),
        }
        for label, expected_value in required.items():
            if label not in label_to_value:
                raise CommandError(f"Stock report EOD block missing '{label}'")
            actual = label_to_value[label]
            if actual is None:
                raise CommandError(f"Stock report EOD block value missing for '{label}'")
            if round(float(actual), 2) != round(float(expected_value), 2):
                raise CommandError(
                    f"Stock report '{label}' mismatch: expected {expected_value}, got {actual}"
                )

        if round(float(eod.x_value), 2) != round(float(required["X Value"]), 2):
            raise CommandError("EOD x_value does not match stock report X Value.")

"""
Django management command to update product stock quantities from products.xlsx.

This script updates stock quantities for existing products based on SKU or prod_code.
It creates InventoryMovement records for audit tracking.

Usage:
    # Update stock from default file (/app/products.xlsx)
    python manage.py update_stock_from_excel

    # Update stock from specific file
    python manage.py update_stock_from_excel --file /path/to/products.xlsx

    # Dry run (preview changes without applying)
    python manage.py update_stock_from_excel --dry-run

    # Add stock column name if different from 'Stock' or 'Quantity'
    python manage.py update_stock_from_excel --stock-column "Current Stock"

Expected Excel Format:
    - Must have columns: Product Name, SKU (or Barcode), and Stock/Quantity column
    - Stock column should contain integer values
    - First row should be headers
"""

from django.core.management.base import BaseCommand
from django.db import transaction
from payments.models import Product, InventoryMovement, User
import openpyxl
import os
from decimal import Decimal


class Command(BaseCommand):
    help = 'Update product stock quantities from Excel file'

    def add_arguments(self, parser):
        parser.add_argument(
            '--file',
            type=str,
            default='/app/products.xlsx',
            help='Path to Excel file (default: /app/products.xlsx)',
        )
        parser.add_argument(
            '--dry-run',
            action='store_true',
            help='Preview changes without applying them',
        )
        parser.add_argument(
            '--stock-column',
            type=str,
            default='Stock',
            help='Name of the stock quantity column (default: Stock)',
        )

    def handle(self, *args, **options):
        excel_file = options['file']
        dry_run = options['dry_run']
        stock_column_name = options['stock_column']

        # Check if file exists
        if not os.path.exists(excel_file):
            self.stdout.write(
                self.style.ERROR(f'Error: File not found: {excel_file}')
            )
            return

        # Read Excel file
        self.stdout.write(f'Reading Excel file: {excel_file}')
        try:
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active  # Get first sheet
        except Exception as e:
            self.stdout.write(
                self.style.ERROR(f'Error reading Excel file: {e}')
            )
            return

        # Parse headers from first row
        headers = []
        for cell in ws[1]:
            headers.append(cell.value)

        self.stdout.write(f'Found columns: {", ".join([h for h in headers if h])}')

        # Find column indices
        try:
            # Try to find stock column (case-insensitive)
            stock_col_idx = None
            for idx, header in enumerate(headers):
                if header and stock_column_name.lower() in header.lower():
                    stock_col_idx = idx
                    break

            # If not found, try common alternatives
            if stock_col_idx is None:
                for idx, header in enumerate(headers):
                    if header and header.lower() in ['stock', 'quantity', 'qty', 'current stock']:
                        stock_col_idx = idx
                        break

            if stock_col_idx is None:
                self.stdout.write(
                    self.style.ERROR(
                        f'Error: Could not find stock column. Available columns: {headers}\n'
                        f'Use --stock-column to specify the correct column name.'
                    )
                )
                return

            # Find product identifier columns (SKU, Barcode, Product Name)
            sku_col_idx = None
            barcode_col_idx = None
            name_col_idx = None

            for idx, header in enumerate(headers):
                if not header:
                    continue
                header_lower = header.lower()
                if 'sku' in header_lower and sku_col_idx is None:
                    sku_col_idx = idx
                elif 'barcode' in header_lower and barcode_col_idx is None:
                    barcode_col_idx = idx
                elif 'product name' in header_lower or header_lower == 'name':
                    name_col_idx = idx

            self.stdout.write(
                self.style.SUCCESS(
                    f'Using stock column: {headers[stock_col_idx]} (index {stock_col_idx})'
                )
            )

        except ValueError as e:
            self.stdout.write(
                self.style.ERROR(f'Error: {e}')
            )
            return

        # Get admin user for audit trail (or create system user)
        admin_user = User.objects.filter(role=User.Role.ADMIN).first()
        if not admin_user:
            admin_user = User.objects.filter(is_superuser=True).first()

        if dry_run:
            self.stdout.write(
                self.style.WARNING('\n🔍 DRY RUN MODE - No changes will be applied\n')
            )

        # Process Excel rows
        updated_count = 0
        created_movements = 0
        not_found_count = 0
        skipped_count = 0
        errors = []

        self.stdout.write('='*80)

        with transaction.atomic():
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not row or all(cell is None for cell in row):
                    continue

                # Extract values
                stock_value = row[stock_col_idx] if stock_col_idx < len(row) else None
                sku_value = row[sku_col_idx] if sku_col_idx is not None and sku_col_idx < len(row) else None
                barcode_value = row[barcode_col_idx] if barcode_col_idx is not None and barcode_col_idx < len(row) else None
                name_value = row[name_col_idx] if name_col_idx is not None and name_col_idx < len(row) else None

                # Skip if no stock value
                if stock_value is None or stock_value == '':
                    continue

                # Parse stock quantity
                try:
                    new_quantity = int(float(stock_value))
                    if new_quantity < 0:
                        errors.append(f'Row {row_idx}: Negative stock value ({new_quantity})')
                        skipped_count += 1
                        continue
                except (ValueError, TypeError):
                    errors.append(f'Row {row_idx}: Invalid stock value ({stock_value})')
                    skipped_count += 1
                    continue

                # Try to find product by SKU, barcode, or name
                product = None
                search_key = None

                # Try SKU first
                if sku_value:
                    sku_str = str(int(sku_value)) if isinstance(sku_value, (int, float)) else str(sku_value)
                    product = Product.objects.filter(sku=sku_str).first()
                    if product:
                        search_key = f"SKU: {sku_str}"

                # Try barcode
                if not product and barcode_value:
                    barcode_str = str(int(barcode_value)) if isinstance(barcode_value, (int, float)) else str(barcode_value)
                    product = Product.objects.filter(barcode=barcode_str).first()
                    if product:
                        search_key = f"Barcode: {barcode_str}"

                # Try prod_code as fallback
                if not product and sku_value:
                    sku_str = str(int(sku_value)) if isinstance(sku_value, (int, float)) else str(sku_value)
                    product = Product.objects.filter(prod_code=sku_str).first()
                    if product:
                        search_key = f"Product Code: {sku_str}"

                # Try product name as last resort
                if not product and name_value:
                    product = Product.objects.filter(prod_name__iexact=name_value).first()
                    if product:
                        search_key = f"Name: {name_value}"

                if not product:
                    identifier = sku_value or barcode_value or name_value or 'Unknown'
                    errors.append(f'Row {row_idx}: Product not found ({identifier})')
                    not_found_count += 1
                    continue

                # Check if stock needs update
                old_quantity = product.quantity
                if old_quantity == new_quantity:
                    self.stdout.write(
                        self.style.WARNING(
                            f'⊘ Row {row_idx}: {product.prod_code} - {product.prod_name} '
                            f'(Stock unchanged: {old_quantity})'
                        )
                    )
                    skipped_count += 1
                    continue

                # Update stock
                quantity_change = new_quantity - old_quantity

                if not dry_run:
                    product.quantity = new_quantity
                    product.save()

                    # Create inventory movement record
                    movement = InventoryMovement.objects.create(
                        product=product,
                        movement_type='ADJUSTMENT',
                        quantity_before=old_quantity,
                        quantity_after=new_quantity,
                        quantity_change=quantity_change,
                        reference=f'Stock update from Excel (Row {row_idx})',
                        performed_by=admin_user.username if admin_user else 'system',
                        notes=f'Updated via update_stock_from_excel command'
                    )
                    created_movements += 1

                # Show update
                sign = '+' if quantity_change > 0 else ''
                color_style = self.style.SUCCESS if quantity_change > 0 else self.style.WARNING

                self.stdout.write(
                    color_style(
                        f'{"[DRY RUN] " if dry_run else ""}✓ Row {row_idx}: '
                        f'{product.prod_code} - {product.prod_name} '
                        f'({search_key}) | Stock: {old_quantity} → {new_quantity} ({sign}{quantity_change})'
                    )
                )
                updated_count += 1

            if dry_run:
                # Rollback transaction in dry-run mode
                transaction.set_rollback(True)

        # Summary
        self.stdout.write(self.style.SUCCESS('\n' + '='*80))
        self.stdout.write(self.style.SUCCESS('UPDATE SUMMARY'))
        self.stdout.write(self.style.SUCCESS('='*80))

        if dry_run:
            self.stdout.write(self.style.WARNING('Mode: DRY RUN (no changes applied)'))

        self.stdout.write(self.style.SUCCESS(f'Updated: {updated_count} products'))
        self.stdout.write(self.style.SUCCESS(f'Inventory Movements Created: {created_movements}'))
        self.stdout.write(self.style.WARNING(f'Unchanged: {skipped_count} products'))
        self.stdout.write(self.style.ERROR(f'Not Found: {not_found_count} products'))

        if errors:
            self.stdout.write(self.style.ERROR(f'\nErrors ({len(errors)}):'))
            for error in errors[:10]:  # Show first 10 errors
                self.stdout.write(self.style.ERROR(f'  • {error}'))
            if len(errors) > 10:
                self.stdout.write(self.style.ERROR(f'  ... and {len(errors) - 10} more'))

        self.stdout.write(self.style.SUCCESS('='*80))

        if dry_run:
            self.stdout.write(
                self.style.WARNING(
                    '\n💡 This was a dry run. Run without --dry-run to apply changes.'
                )
            )
        else:
            self.stdout.write(
                self.style.SUCCESS(
                    f'\n✓ Stock update completed! Updated {updated_count} products.'
                )
            )

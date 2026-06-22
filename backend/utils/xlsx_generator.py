import logging
from io import BytesIO
from typing import Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

THIN_BORDER = Border(
    left=Side(style='thin', color='CBD5E0'),
    right=Side(style='thin', color='CBD5E0'),
    top=Side(style='thin', color='CBD5E0'),
    bottom=Side(style='thin', color='CBD5E0'),
)
HEADER_FONT = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
HEADER_FILL = PatternFill(start_color='4A5568', end_color='4A5568', fill_type='solid')
HEADER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
TITLE_FONT = Font(name='Calibri', size=13, bold=True, color='0891B2')
NUM_ALIGN = Alignment(horizontal='right', vertical='center')
LEFT_ALIGN = Alignment(horizontal='left', vertical='center')
CENTER_ALIGN = Alignment(horizontal='center', vertical='center')
DATA_FONT = Font(name='Calibri', size=10)
STRIPE_FILL = PatternFill(start_color='F7FAFC', end_color='F7FAFC', fill_type='solid')


class XlsxGenerator:

    @staticmethod
    def from_data(
        data: List[Dict],
        columns: List[Dict],
        sheet_name: str = 'Sheet1',
        title: Optional[str] = None,
    ) -> BytesIO:
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name[:31]
        XlsxGenerator._populate_sheet(ws, data, columns, title)
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    @staticmethod
    def multi_sheet(sheets: List[Dict]) -> BytesIO:
        wb = Workbook()
        first = True
        for sheet in sheets:
            if first:
                ws = wb.active
                ws.title = (sheet.get('name', 'Sheet1')[:31])
                first = False
            else:
                ws = wb.create_sheet(title=sheet.get('name', f'Sheet{sheets.index(sheet) + 1}')[:31])
            XlsxGenerator._populate_sheet(
                ws,
                sheet.get('data', []),
                sheet.get('columns', []),
                sheet.get('title'),
            )
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    @staticmethod
    def _populate_sheet(ws, data: List[Dict], columns: List[Dict], title: Optional[str] = None):
        row = 1
        if title:
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
            cell = ws.cell(row=1, column=1, value=title)
            cell.font = TITLE_FONT
            cell.alignment = CENTER_ALIGN
            row = 2

        header_row = row
        for col, col_def in enumerate(columns, 1):
            cell = ws.cell(row=row, column=col, value=col_def.get('header', ''))
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGN
            cell.border = THIN_BORDER
            width = col_def.get('width', 15)
            ws.column_dimensions[get_column_letter(col)].width = width
        row += 1

        for i, item in enumerate(data):
            stripe = i % 2 == 1
            for col, col_def in enumerate(columns, 1):
                key = col_def.get('key', '')
                val = item.get(key, '')
                cell = ws.cell(row=row, column=col, value=val)
                cell.font = DATA_FONT
                cell.border = THIN_BORDER
                if col_def.get('align') == 'right':
                    cell.alignment = NUM_ALIGN
                elif col_def.get('align') == 'center':
                    cell.alignment = CENTER_ALIGN
                else:
                    cell.alignment = LEFT_ALIGN
                if stripe:
                    cell.fill = STRIPE_FILL
            row += 1

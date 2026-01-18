#!/usr/bin/env python
"""
Production Setup Script for Go-Live on 2026-01-19

This script prepares the system for production use by:
1. Verifying unfulfilled transactions match Excel
2. Setting up revenue reconciliation baseline
3. Preparing inventory baseline (to be run after physical count)

Usage:
    # Step 1: Verify unfulfilled transactions
    docker exec -it inventory-management-web-1 python scripts/production_setup.py verify

    # Step 2: After updating Excel with today's unfulfilled, run again
    docker exec -it inventory-management-web-1 python scripts/production_setup.py verify

    # Step 3: After physical stock count, update inventory
    docker exec -it inventory-management-web-1 python scripts/production_setup.py set_inventory

    # Step 4: Generate baseline report
    docker exec -it inventory-management-web-1 python scripts/production_setup.py baseline_report
"""

import os
import sys
from datetime import date, datetime, timedelta
from decimal import Decimal

# Django setup
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'management.settings')

import django
django.setup()

from django.db import transaction as db_transaction
from django.db.models import Sum, Count, Q
from django.utils import timezone
from payments.models import Transaction, PaymentGateway, Product, InventoryMovement

# Configuration
LAUNCH_DATE = date(2026, 1, 19)  # Tomorrow - first day of production use
EXCEL_PATH = os.path.join(os.path.dirname(__file__), '..', 'unused.xlsx')


def parse_unfulfilled_excel(excel_path):
    """Parse unfulfilled transactions from Excel file.

    Expected columns: TX ID, Amount, Date
    """
    import openpyxl

    if not os.path.exists(excel_path):
        print(f"ERROR: Excel file not found at {excel_path}")
        return []

    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb.active

    unfulfilled = []

    # Skip header row, start from row 2
    for row_num in range(2, ws.max_row + 1):
        row = [cell.value for cell in ws[row_num]]

        tx_id = row[0]  # Column A: TX ID
        amount = row[1]  # Column B: Amount
        tx_date = row[2] if len(row) > 2 else None  # Column C: Date

        if tx_id and amount:
            try:
                # Skip formula cells or header-like rows
                if str(amount).startswith('=') or str(tx_id).upper() == 'TX ID':
                    continue
                amt = Decimal(str(amount))
                unfulfilled.append({
                    'date': tx_date,
                    'tx_id': str(tx_id).strip(),
                    'amount': amt
                })
            except:
                pass

    return unfulfilled


def verify_unfulfilled():
    """
    Compare Excel unfulfilled list against system transactions.
    Reports discrepancies.
    """
    print("=" * 60)
    print("VERIFYING UNFULFILLED TRANSACTIONS")
    print("=" * 60)

    # Parse Excel
    excel_unfulfilled = parse_unfulfilled_excel(EXCEL_PATH)
    print(f"\nExcel unfulfilled: {len(excel_unfulfilled)} transactions")
    excel_total = sum(t['amount'] for t in excel_unfulfilled)
    print(f"Excel total: KES {excel_total:,.2f}")

    # Get system unfulfilled (NOT_PROCESSED and PROCESSING on Paybill)
    paybill = PaymentGateway.objects.filter(
        gateway_type=PaymentGateway.GatewayType.MPESA_PAYBILL,
        is_parent_company=True
    ).first()

    if not paybill:
        paybill = PaymentGateway.objects.filter(
            gateway_type=PaymentGateway.GatewayType.MPESA_PAYBILL
        ).first()

    system_unfulfilled = Transaction.objects.filter(
        gateway=paybill,
        status__in=['NOT_PROCESSED', 'PROCESSING']
    ).exclude(
        status='COMBINED_FULFILLED'
    )

    system_count = system_unfulfilled.count()
    system_total = system_unfulfilled.aggregate(total=Sum('amount'))['total'] or Decimal('0')

    print(f"\nSystem unfulfilled (Paybill): {system_count} transactions")
    print(f"System total: KES {system_total:,.2f}")

    # Build lookup sets
    excel_tx_ids = {t['tx_id'] for t in excel_unfulfilled}
    system_tx_ids = set(system_unfulfilled.values_list('tx_id', flat=True))

    # Find discrepancies
    in_excel_not_system = excel_tx_ids - system_tx_ids
    in_system_not_excel = system_tx_ids - excel_tx_ids

    print("\n" + "-" * 40)
    print("DISCREPANCIES")
    print("-" * 40)

    if in_excel_not_system:
        print(f"\n❌ In Excel but NOT in system ({len(in_excel_not_system)}):")
        for tx_id in sorted(in_excel_not_system):
            excel_tx = next(t for t in excel_unfulfilled if t['tx_id'] == tx_id)
            print(f"   {tx_id}: KES {excel_tx['amount']:,.2f}")

            # Check if it exists with different status
            existing = Transaction.objects.filter(tx_id=tx_id).first()
            if existing:
                print(f"      → EXISTS in system with status: {existing.status}")
    else:
        print("\n✓ All Excel transactions found in system")

    if in_system_not_excel:
        print(f"\n⚠️  In system but NOT in Excel ({len(in_system_not_excel)}):")
        for tx_id in sorted(list(in_system_not_excel)[:20]):  # Limit to 20
            txn = system_unfulfilled.get(tx_id=tx_id)
            print(f"   {tx_id}: KES {txn.amount:,.2f} | {txn.sender_name}")
        if len(in_system_not_excel) > 20:
            print(f"   ... and {len(in_system_not_excel) - 20} more")
    else:
        print("\n✓ All system unfulfilled transactions are in Excel")

    # Check for amount mismatches
    print("\n" + "-" * 40)
    print("AMOUNT MISMATCHES")
    print("-" * 40)

    mismatches = []
    for excel_tx in excel_unfulfilled:
        if excel_tx['tx_id'] in system_tx_ids:
            system_tx = system_unfulfilled.get(tx_id=excel_tx['tx_id'])
            if system_tx.amount != excel_tx['amount']:
                mismatches.append({
                    'tx_id': excel_tx['tx_id'],
                    'excel_amount': excel_tx['amount'],
                    'system_amount': system_tx.amount
                })

    if mismatches:
        print(f"\n❌ Amount mismatches ({len(mismatches)}):")
        for m in mismatches:
            print(f"   {m['tx_id']}: Excel={m['excel_amount']:,.2f}, System={m['system_amount']:,.2f}")
    else:
        print("\n✓ All matching transactions have correct amounts")

    # Summary
    print("\n" + "=" * 60)
    print("SUMMARY")
    print("=" * 60)
    print(f"Excel transactions: {len(excel_unfulfilled)}")
    print(f"System unfulfilled: {system_count}")
    print(f"In Excel only: {len(in_excel_not_system)}")
    print(f"In System only: {len(in_system_not_excel)}")
    print(f"Amount mismatches: {len(mismatches)}")

    if not in_excel_not_system and not mismatches:
        print("\n✅ VERIFICATION PASSED - Excel matches system")
    else:
        print("\n⚠️  VERIFICATION NEEDS ATTENTION - Review discrepancies above")

    return {
        'excel_count': len(excel_unfulfilled),
        'excel_total': excel_total,
        'system_count': system_count,
        'system_total': system_total,
        'in_excel_not_system': in_excel_not_system,
        'in_system_not_excel': in_system_not_excel,
        'mismatches': mismatches
    }


def set_inventory_baseline():
    """
    Set inventory baseline from user input.
    Run this after physical stock count.
    """
    print("=" * 60)
    print("SET INVENTORY BASELINE")
    print("=" * 60)

    products = Product.objects.filter(is_active=True).order_by('prod_name')

    print(f"\nFound {products.count()} active products")
    print("\nCurrent inventory levels:")
    print("-" * 50)

    for p in products:
        print(f"  {p.prod_code}: {p.prod_name} = {p.quantity}")

    print("\n" + "-" * 50)
    print("To update inventory, edit this script or use Django admin.")
    print("After physical count, update each product's quantity.")
    print("-" * 50)

    # Example update format (uncomment and modify after physical count)
    """
    inventory_updates = {
        'REG_KIT_001': 50,      # Registration Kit
        '4IN1CORDYC': 100,      # 4 in 1 Cordyceps Coffee
        # Add all products here
    }

    with db_transaction.atomic():
        for prod_code, new_quantity in inventory_updates.items():
            try:
                product = Product.objects.select_for_update().get(prod_code=prod_code)
                old_quantity = product.quantity
                product.quantity = new_quantity
                product.save()

                # Create audit record
                InventoryMovement.objects.create(
                    movement_type=InventoryMovement.MovementType.ADJUSTMENT,
                    product=product,
                    quantity_before=old_quantity,
                    quantity_after=new_quantity,
                    quantity_change=new_quantity - old_quantity,
                    reference=f'Production baseline setup {date.today()}',
                    performed_by='System'
                )
                print(f"Updated {prod_code}: {old_quantity} → {new_quantity}")
            except Product.DoesNotExist:
                print(f"WARNING: Product {prod_code} not found")
    """


def generate_baseline_report():
    """
    Generate a baseline report of current system state.
    Save this for reference.
    """
    print("=" * 60)
    print(f"BASELINE REPORT - {date.today()}")
    print("=" * 60)

    # Transaction summary
    print("\n--- TRANSACTIONS ---")
    statuses = Transaction.objects.exclude(
        status='COMBINED_FULFILLED'
    ).values('status').annotate(
        count=Count('id'),
        total=Sum('amount'),
        fulfilled=Sum('amount_fulfilled')
    )

    for s in statuses:
        print(f"  {s['status']}: {s['count']} txns, Amount: KES {s['total'] or 0:,.2f}, Fulfilled: KES {s['fulfilled'] or 0:,.2f}")

    # Unfulfilled by gateway
    print("\n--- UNFULFILLED BY GATEWAY ---")
    unfulfilled = Transaction.objects.filter(
        status__in=['NOT_PROCESSED', 'PROCESSING']
    ).exclude(
        status='COMBINED_FULFILLED'
    ).values('gateway__name').annotate(
        count=Count('id'),
        total=Sum('amount')
    )

    for u in unfulfilled:
        gateway = u['gateway__name'] or 'No Gateway'
        print(f"  {gateway}: {u['count']} txns, KES {u['total'] or 0:,.2f}")

    # Partially fulfilled
    print("\n--- PARTIALLY FULFILLED ---")
    partial = Transaction.objects.filter(status='PARTIALLY_FULFILLED')
    partial_data = []
    for p in partial:
        remaining = p.amount - p.amount_fulfilled
        partial_data.append({
            'tx_id': p.tx_id,
            'amount': p.amount,
            'fulfilled': p.amount_fulfilled,
            'remaining': remaining
        })

    total_remaining = sum(d['remaining'] for d in partial_data)
    print(f"  Count: {len(partial_data)}")
    print(f"  Total remaining: KES {total_remaining:,.2f}")

    # Inventory
    print("\n--- INVENTORY ---")
    products = Product.objects.filter(is_active=True).order_by('prod_name')
    for p in products:
        print(f"  {p.prod_code}: {p.quantity} units")

    # Save to file
    report_path = os.path.join(
        os.path.dirname(__file__),
        f'baseline_report_{date.today().isoformat()}.txt'
    )

    with open(report_path, 'w') as f:
        f.write(f"BASELINE REPORT - {date.today()}\n")
        f.write("=" * 60 + "\n\n")
        f.write("This is the system state before production launch.\n\n")

        f.write("TRANSACTIONS:\n")
        for s in statuses:
            f.write(f"  {s['status']}: {s['count']} txns, KES {s['total'] or 0:,.2f}\n")

        f.write("\nUNFULFILLED:\n")
        for u in unfulfilled:
            gateway = u['gateway__name'] or 'No Gateway'
            f.write(f"  {gateway}: {u['count']} txns, KES {u['total'] or 0:,.2f}\n")

        f.write(f"\nPARTIALLY FULFILLED REMAINING: KES {total_remaining:,.2f}\n")

        f.write("\nINVENTORY:\n")
        for p in products:
            f.write(f"  {p.prod_code}: {p.quantity} units\n")

    print(f"\n✅ Report saved to: {report_path}")


def fix_manual_payment_gateways():
    """
    Assign correct gateways to manual payments (MAN-PDQ, MAN-CAS, etc).
    """
    print("=" * 60)
    print("FIX MANUAL PAYMENT GATEWAYS")
    print("=" * 60)

    # Get gateways
    pdq_gateway = PaymentGateway.objects.filter(gateway_type='PDQ').first()
    cash_gateway = PaymentGateway.objects.filter(gateway_type='CASH').first()
    bank_gateway = PaymentGateway.objects.filter(gateway_type='BANK_TRANSFER').first()

    print(f"PDQ Gateway: {pdq_gateway}")
    print(f"Cash Gateway: {cash_gateway}")
    print(f"Bank Gateway: {bank_gateway}")

    # Find manual payments without gateway
    manual_no_gateway = Transaction.objects.filter(
        tx_id__startswith='MAN-',
        gateway__isnull=True
    )

    print(f"\nManual payments without gateway: {manual_no_gateway.count()}")

    # Categorize by prefix
    pdq_txns = manual_no_gateway.filter(tx_id__startswith='MAN-PDQ')
    cash_txns = manual_no_gateway.filter(tx_id__startswith='MAN-CAS')
    bank_txns = manual_no_gateway.filter(tx_id__startswith='MAN-BNK')

    print(f"  MAN-PDQ: {pdq_txns.count()}")
    print(f"  MAN-CAS: {cash_txns.count()}")
    print(f"  MAN-BNK: {bank_txns.count()}")

    confirm = input("\nAssign gateways to these transactions? (yes/no): ")
    if confirm.lower() != 'yes':
        print("Cancelled.")
        return

    with db_transaction.atomic():
        if pdq_gateway:
            updated = pdq_txns.update(gateway=pdq_gateway)
            print(f"✓ Assigned PDQ gateway to {updated} transactions")

        if cash_gateway:
            updated = cash_txns.update(gateway=cash_gateway)
            print(f"✓ Assigned Cash gateway to {updated} transactions")

        if bank_gateway:
            updated = bank_txns.update(gateway=bank_gateway)
            print(f"✓ Assigned Bank gateway to {updated} transactions")

    print("\n✅ Gateway assignment complete")


def main():
    if len(sys.argv) < 2:
        print(__doc__)
        print("\nAvailable commands:")
        print("  verify          - Verify unfulfilled transactions against Excel")
        print("  set_inventory   - Set inventory baseline (after physical count)")
        print("  baseline_report - Generate baseline report")
        print("  fix_gateways    - Assign gateways to manual payments")
        return

    command = sys.argv[1]

    if command == 'verify':
        verify_unfulfilled()
    elif command == 'set_inventory':
        set_inventory_baseline()
    elif command == 'baseline_report':
        generate_baseline_report()
    elif command == 'fix_gateways':
        fix_manual_payment_gateways()
    else:
        print(f"Unknown command: {command}")
        print("Run without arguments for help.")


if __name__ == '__main__':
    main()

#!/usr/bin/env python3
"""
Quick script to inspect Excel file structure before running stock update.

Usage:
    python inspect_excel.py products.xlsx
"""

import sys
import openpyxl

def inspect_excel(file_path):
    """Inspect Excel file and show structure"""
    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active

        print(f"\n{'='*80}")
        print(f"EXCEL FILE INSPECTION: {file_path}")
        print(f"{'='*80}\n")

        # Get headers
        headers = [cell.value for cell in ws[1]]
        print(f"📋 HEADERS ({len([h for h in headers if h])} columns):")
        for idx, header in enumerate(headers):
            if header:
                print(f"  [{idx}] {header}")

        # Count rows
        row_count = sum(1 for row in ws.iter_rows(min_row=2, values_only=True) if any(cell for cell in row))
        print(f"\n📊 DATA ROWS: {row_count}")

        # Show first 5 data rows as sample
        print(f"\n🔍 SAMPLE DATA (first 5 rows):")
        print("-" * 80)

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=6, values_only=True), start=2):
            if not any(row):
                continue
            print(f"\nRow {row_idx}:")
            for idx, (header, value) in enumerate(zip(headers, row)):
                if header and value:
                    print(f"  {header}: {value}")

        # Check for stock column
        print(f"\n{'='*80}")
        print("🔍 STOCK COLUMN DETECTION:")
        stock_keywords = ['stock', 'quantity', 'qty', 'current stock']
        stock_cols = []
        for idx, header in enumerate(headers):
            if header and any(keyword in header.lower() for keyword in stock_keywords):
                stock_cols.append((idx, header))

        if stock_cols:
            print("✓ Found potential stock columns:")
            for idx, header in stock_cols:
                # Count non-empty values in this column
                values = [row[idx] for row in ws.iter_rows(min_row=2, values_only=True) if idx < len(row) and row[idx] is not None]
                print(f"  [{idx}] {header} ({len(values)} non-empty values)")
        else:
            print("✗ No stock column found! You may need to specify --stock-column")

        # Check for product identifier columns
        print(f"\n🔍 PRODUCT IDENTIFIER COLUMNS:")
        id_keywords = {
            'SKU': ['sku'],
            'Barcode': ['barcode'],
            'Product Name': ['product name', 'name', 'prod_name'],
            'Product Code': ['product code', 'prod_code', 'code']
        }

        found_identifiers = []
        for id_type, keywords in id_keywords.items():
            for idx, header in enumerate(headers):
                if header and any(keyword in header.lower() for keyword in keywords):
                    found_identifiers.append((id_type, idx, header))
                    break

        if found_identifiers:
            print("✓ Found product identifier columns:")
            for id_type, idx, header in found_identifiers:
                # Count non-empty values
                values = [row[idx] for row in ws.iter_rows(min_row=2, values_only=True) if idx < len(row) and row[idx]]
                print(f"  {id_type}: [{idx}] {header} ({len(values)} non-empty values)")
        else:
            print("✗ No product identifier columns found!")

        print(f"\n{'='*80}")
        print("✓ Inspection complete!")
        print(f"{'='*80}\n")

        # Recommendations
        print("💡 RECOMMENDATIONS:")
        if not stock_cols:
            print("  ⚠ Add --stock-column parameter with your stock column name")
        if not found_identifiers:
            print("  ⚠ Ensure your Excel has SKU, Barcode, or Product Name column")
        if stock_cols and found_identifiers:
            print("  ✓ Excel file looks good! Ready to run update_stock_from_excel")
            print("  ✓ Remember to use --dry-run first!")

        print()

    except FileNotFoundError:
        print(f"❌ Error: File not found: {file_path}")
        sys.exit(1)
    except Exception as e:
        print(f"❌ Error reading file: {e}")
        sys.exit(1)


if __name__ == '__main__':
    if len(sys.argv) != 2:
        print("Usage: python inspect_excel.py <excel_file>")
        print("Example: python inspect_excel.py products.xlsx")
        sys.exit(1)

    inspect_excel(sys.argv[1])

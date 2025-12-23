#!/usr/bin/env python
"""Read products from Excel file and display structure."""

import openpyxl
import sys

def read_excel(filepath):
    """Read Excel file and display contents."""
    wb = openpyxl.load_workbook(filepath)

    print("Available sheets:", wb.sheetnames)
    print()

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n=== Sheet: {sheet_name} ===")

        # Get headers (first row)
        headers = []
        for cell in ws[1]:
            headers.append(cell.value)
        print(f"Headers: {headers}")
        print()

        # Display all rows
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, values_only=True), start=1):
            if row_idx == 1:
                print(f"Row {row_idx} (Headers): {row}")
            else:
                print(f"Row {row_idx}: {row}")

        print(f"\nTotal rows: {ws.max_row}")
        print(f"Total columns: {ws.max_column}")

if __name__ == "__main__":
    filepath = "/app/products.xlsx"
    read_excel(filepath)
